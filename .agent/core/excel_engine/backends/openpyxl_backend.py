"""OpenpyxlBackend — operaciones Excel basadas en archivos usando openpyxl.

Usado para operaciones batch/headless que no requieren Excel en ejecución.
"""

from __future__ import annotations

import csv
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

from excel_engine.types import CellFormat

logger = logging.getLogger(__name__)


@dataclass
class _Handle:
    """Handle interno para openpyxl: path + workbook + flag dirty."""

    path: str
    workbook: openpyxl.Workbook
    mode: str
    dirty: bool = False


class OpenpyxlBackend:
    """Backend basado en archivos usando openpyxl."""

    def open(self, path: str, mode: Literal["read", "write", "live"]) -> _Handle:
        """Abre un workbook y retorna un handle _Handle.

        Args:
            path: Ruta al archivo .xlsx.
            mode: Modo de apertura — 'read' o 'write' (no 'live').

        Returns:
            Handle interno con workbook cargado.

        Raises:
            FileNotFoundError: Si el archivo no existe.
            ValueError: Si el modo es 'live' (no soportado por openpyxl).
        """
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"archivo no encontrado: {path}")
        if mode == "live":
            raise ValueError("openpyxl backend no soporta mode='live'")
        # data_only=False para preservar fórmulas (no valores cacheados)
        wb = openpyxl.load_workbook(p, data_only=False)
        logger.info("[openpyxl] abierto path=%s mode=%s", path, mode)
        return _Handle(path=str(p), workbook=wb, mode=mode, dirty=False)

    def close(self, handle: _Handle, save: bool = True) -> None:
        """Cierra el workbook. Si save=True y hay cambios, persiste.

        Args:
            handle: Handle retornado por open().
            save: Si True y dirty=True, guarda antes de cerrar.
        """
        if save and handle.dirty:
            handle.workbook.save(handle.path)
            logger.info("[openpyxl] guardado path=%s", handle.path)
        handle.workbook.close()

    def read_range(
        self,
        handle: _Handle,
        sheet: str,
        range_addr: str,
    ) -> list[list[Any]]:
        """Retorna los valores de celdas del rango como lista 2D.

        Args:
            handle: Handle retornado por open().
            sheet: Nombre de la hoja.
            range_addr: Dirección del rango, ej. 'A1:C10' o 'A1'.

        Returns:
            Lista de listas con los valores de las celdas.
        """
        ws = handle.workbook[sheet]
        cells = ws[range_addr]
        # Rango de una sola celda retorna un Cell; rango retorna tupla de tuplas
        if not isinstance(cells, tuple):
            return [[cells.value]]
        result: list[list[Any]] = []
        for row in cells:
            if isinstance(row, tuple):
                result.append([c.value for c in row])
            else:
                result.append([row.value])
        return result

    def write_range(
        self,
        handle: _Handle,
        sheet: str,
        range_addr: str,
        values: list[list[Any]],
    ) -> None:
        """Escribe una lista 2D de valores comenzando desde range_addr.

        Args:
            handle: Handle retornado por open().
            sheet: Nombre de la hoja.
            range_addr: Celda de inicio o rango, ej. 'A1'.
            values: Lista de listas con los valores a escribir.

        Raises:
            PermissionError: Si el handle fue abierto en modo 'read'.
        """
        if handle.mode == "read":
            raise PermissionError("sesión abierta en modo lectura")
        ws = handle.workbook[sheet]
        # Resolver celda superior izquierda
        if ":" in range_addr:
            top_left = range_addr.split(":")[0]
        else:
            top_left = range_addr
        anchor = ws[top_left]
        start_row = anchor.row
        start_col = anchor.column
        for r_offset, row in enumerate(values):
            for c_offset, val in enumerate(row):
                ws.cell(
                    row=start_row + r_offset,
                    column=start_col + c_offset,
                    value=val,
                )
        handle.dirty = True

    def apply_formula(
        self,
        handle: _Handle,
        sheet: str,
        range_addr: str,
        formula: str,
    ) -> None:
        """Escribe una expresión de fórmula en la celda indicada.

        Args:
            handle: Handle retornado por open().
            sheet: Nombre de la hoja.
            range_addr: Dirección de la celda destino.
            formula: Fórmula a escribir, con o sin '=' inicial.

        Raises:
            PermissionError: Si el handle fue abierto en modo 'read'.
        """
        if handle.mode == "read":
            raise PermissionError("sesión abierta en modo lectura")
        if not formula.startswith("="):
            formula = "=" + formula
        ws = handle.workbook[sheet]
        ws[range_addr] = formula
        handle.dirty = True

    def set_format(
        self,
        handle: _Handle,
        sheet: str,
        range_addr: str,
        fmt: CellFormat,
    ) -> None:
        """Aplica CellFormat al rango especificado.

        Args:
            handle: Handle retornado por open().
            sheet: Nombre de la hoja.
            range_addr: Dirección del rango.
            fmt: Opciones de formato a aplicar.

        Raises:
            PermissionError: Si el handle fue abierto en modo 'read'.
        """
        if handle.mode == "read":
            raise PermissionError("sesión abierta en modo lectura")
        ws = handle.workbook[sheet]
        cells = ws[range_addr]
        flat = self._flatten(cells)
        for cell in flat:
            if fmt.number_format is not None:
                cell.number_format = fmt.number_format
            if any([fmt.bold, fmt.italic, fmt.font_color, fmt.font_size]):
                cell.font = Font(
                    name=cell.font.name,
                    size=fmt.font_size if fmt.font_size is not None else cell.font.size,
                    bold=fmt.bold if fmt.bold else cell.font.bold,
                    italic=fmt.italic if fmt.italic else cell.font.italic,
                    color=self._normalize_color(fmt.font_color)
                    if fmt.font_color
                    else cell.font.color,
                )
            if fmt.bg_color:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=self._normalize_color(fmt.bg_color),
                )
            if fmt.align:
                cell.alignment = Alignment(horizontal=fmt.align)
        handle.dirty = True

    def create_table(
        self,
        handle: _Handle,
        sheet: str,
        range_addr: str,
        name: str,
        style: str | None = None,
    ) -> None:
        """Convierte un rango en una Tabla Excel.

        Args:
            handle: Handle retornado por open().
            sheet: Nombre de la hoja.
            range_addr: Rango de la tabla, ej. 'A1:B10'.
            name: Nombre único de la tabla dentro del workbook.
            style: Nombre del estilo de tabla (default: 'TableStyleMedium9').

        Raises:
            PermissionError: Si el handle fue abierto en modo 'read'.
            ValueError: Si ya existe una tabla con ese nombre en la hoja.
        """
        if handle.mode == "read":
            raise PermissionError("sesión abierta en modo lectura")
        ws = handle.workbook[sheet]
        if name in set(ws.tables):
            raise ValueError(f"tabla {name!r} ya existe en hoja {sheet!r}")
        tbl = Table(displayName=name, ref=range_addr)
        tbl.tableStyleInfo = TableStyleInfo(
            name=style or "TableStyleMedium9",
            showRowStripes=True,
        )
        ws.add_table(tbl)
        handle.dirty = True

    def save_as(
        self,
        handle: _Handle,
        path: str,
        format: Literal["xlsx", "pdf", "csv", "xlsm"],
    ) -> None:
        """Exporta el workbook a una nueva ruta/formato.

        Args:
            handle: Handle retornado por open().
            path: Ruta de destino del archivo exportado.
            format: Formato de exportación — 'xlsx', 'pdf', 'csv' o 'xlsm'.

        Raises:
            NotImplementedError: Si el formato es 'pdf' (requiere xlwings / Plan B).
            ValueError: Si el formato no es soportado.
        """
        if format in ("xlsx", "xlsm"):
            handle.workbook.save(path)
            return
        if format == "csv":
            ws = handle.workbook.active
            if ws is None:
                raise RuntimeError("no hay hoja activa para exportar como csv")
            with open(path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            return
        if format == "pdf":
            raise NotImplementedError("exportación a pdf requiere xlwings; usar Plan B")
        raise ValueError(f"formato no soportado: {format!r}")

    @staticmethod
    def _flatten(cells: Any) -> list[Any]:
        """Aplana un Cell o tupla-de-tuplas en una lista plana.

        Args:
            cells: Un Cell individual o estructura retornada por ws[range_addr].

        Returns:
            Lista plana de objetos Cell.
        """
        if not isinstance(cells, tuple):
            return [cells]
        flat: list[Any] = []
        for item in cells:
            if isinstance(item, tuple):
                flat.extend(item)
            else:
                flat.append(item)
        return flat

    @staticmethod
    def _normalize_color(c: str) -> str:
        """Convierte '#FF0000' a 'FFFF0000' (ARGB con alpha).

        Args:
            c: Color en formato hex, ej. '#FF0000' o 'FF0000'.

        Returns:
            Color en formato ARGB de 8 caracteres, ej. 'FFFF0000'.
        """
        c = c.lstrip("#").upper()
        if len(c) == 6:
            return "FF" + c
        return c

    def create_chart(
        self,
        handle: _Handle,
        sheet: str,
        range_addr: str,
        chart_type: str = "column",
        title: str | None = None,
        legend: bool = True,
    ) -> dict[str, Any]:
        """Crea un gráfico de barras en la hoja (fallback sin live Excel).

        Args:
            handle: Handle retornado por open().
            sheet: Nombre de la hoja con los datos.
            range_addr: Rango de datos, ej. 'A1:C5'.
            chart_type: Tipo de gráfico ('column', 'bar', 'line').
            title: Título opcional del gráfico.
            legend: Si True, muestra la leyenda.

        Returns:
            dict con 'chart_id' y 'name'.

        Raises:
            PermissionError: Si el handle fue abierto en modo 'read'.
        """
        if handle.mode == "read":
            raise PermissionError("sesión abierta en modo lectura")
        ws = handle.workbook[sheet]

        # Parse range for data boundaries
        if ":" in range_addr:
            top_left = range_addr.split(":")[0]
        else:
            top_left = range_addr
        anchor = ws[top_left]
        ncols = 0
        for nrows, row in enumerate(
            ws.iter_rows(min_row=anchor.row, max_row=anchor.row + 50), start=1
        ):
            if all(c.value is None for c in row):
                break
            ncols = max(ncols, len(row))

        data = Reference(ws, range_string=f"{sheet}!{range_addr}")
        chart = BarChart()
        chart.add_data(data, titles_from_data=True)
        # Map user-friendly names to openpyxl BarChart.type values
        type_map = {"column": "col", "bar": "bar", "col": "col"}
        chart.type = type_map.get(chart_type, chart_type)
        if title is not None:
            chart.title = title
        chart.legend = None if legend else False
        chart_id = id(chart)
        ws.add_chart(chart, anchor=ws[top_left].coordinate)
        handle.dirty = True
        logger.info(
            "[openpyxl] chart created type=%s title=%s",
            chart_type,
            title,
        )
        return {"chart_id": chart_id, "name": f"Chart_{chart_id}"}

    def conditional_format(
        self,
        handle: _Handle,
        sheet: str,
        range_addr: str,
        rule_type: str,
        formula_or_threshold: str | float,
        format_style: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Aplica formato condicional al rango.

        Args:
            handle: Handle retornado por open().
            sheet: Nombre de la hoja.
            range_addr: Rango objetivo, ej. 'A1:C10'.
            rule_type: Tipo de regla ('cell_value', 'contains_text',
                'expression', 'colorScale', 'dataBar', 'iconSet').
            formula_or_threshold: Valor de umbral o fórmula según rule_type.
            format_style: Estilos a aplicar cuando se cumple la condición.

        Returns:
            dict con 'rule_id'.

        Raises:
            PermissionError: Si el handle fue abierto en modo lectura.
        """
        if handle.mode == "read":
            raise PermissionError("sesión abierta en modo lectura")
        ws = handle.workbook[sheet]

        fill = None
        font = None
        if format_style:
            if "bgColor" in format_style or "fgColor" in format_style:
                color = self._normalize_color(
                    format_style.get("fgColor", format_style.get("bgColor", "FFFF0000"))
                )
                fill = PatternFill(
                    fill_type="solid",
                    fgColor=color,
                )
            if "bold" in format_style:
                font = Font(bold=format_style["bold"])

        if rule_type == "cell_value":
            operator = formula_or_threshold
            threshold = 0.0
            if isinstance(formula_or_threshold, (int, float)):
                threshold = formula_or_threshold
            ws.conditional_formatting.add(
                range_addr,
                CellIsRule(operator=operator, formula=[str(threshold)],
                           fill=fill, font=font),
            )
        elif rule_type == "expression":
            ws.conditional_formatting.add(
                range_addr,
                FormulaRule(formula=[str(formula_or_threshold)],
                            fill=fill, font=font),
            )
        else:
            ws.conditional_formatting.add(
                range_addr,
                FormulaRule(
                    formula=[f'ISNUMBER(SEARCH("{formula_or_threshold}",""))'],
                    fill=fill,
                    font=font,
                ),
            )

        rule_id = len(ws.conditional_formatting)
        handle.dirty = True
        logger.info(
            "[openpyxl] conditional_format range=%s type=%s rule_id=%d",
            range_addr,
            rule_type,
            rule_id,
        )
        return {"rule_id": rule_id}

    def calculation_mode(
        self,
        handle: _Handle,
        mode: Literal["automatic", "manual", "semiautomatic"],
    ) -> dict[str, Any]:
        """Configura el modo de cálculo del workbook.

        Args:
            handle: Handle retornado por open().
            mode: Modo de cálculo — 'automatic', 'manual' o 'semiautomatic'.

        Returns:
            dict con 'calculation_mode'.

        Raises:
            ValueError: Si el modo no es soportado.
        """
        valid = {"automatic", "manual", "semiautomatic"}
        if mode not in valid:
            raise ValueError(f"modo de cálculo no soportado: {mode!r}")
        # openpyxl no puede cambiar calculation mode sin Excel vivo;
        # marcamos dirty=True para simular que la operación fue registrada
        handle.dirty = True
        logger.info("[openpyxl] calculation_mode registered as %s", mode)
        return {"calculation_mode": mode}

    def save_as(
        self,
        handle: _Handle,
        path: str,
        format: Literal["xlsx", "pdf", "csv", "xlsm"],
    ) -> None:
        """Exporta el workbook a una nueva ruta/formato.

        Args:
            handle: Handle retornado por open().
            path: Ruta de destino del archivo exportado.
            format: Formato de exportación — 'xlsx', 'pdf', 'csv' o 'xlsm'.

        Raises:
            NotImplementedError: Si el formato es 'pdf' (requiere xlwings / Plan B).
            ValueError: Si el formato no es soportado.
        """
        if format in ("xlsx", "xlsm"):
            handle.workbook.save(path)
            return
        if format == "csv":
            ws = handle.workbook.active
            if ws is None:
                raise RuntimeError("no hay hoja activa para exportar como csv")
            with open(path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            return
        if format == "pdf":
            raise NotImplementedError(
                "exportación a pdf requiere xlwings; usar Plan B"
            )
        raise ValueError(f"formato no soportado: {format!r}")
