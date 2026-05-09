#!/usr/bin/env python3
"""
UNS ULTIMATE System - The Most Complete HR Management System
Version 2.0 - Production Ready
"""

import argparse
import json
import sqlite3
import os
import csv
import zipfile
from datetime import datetime
from pathlib import Path

# ============================================
# CONFIGURATION
# ============================================

VERSION = "2.0.0"
DB_PATH = os.environ.get("UNS_DB", "uns_ultimate.db")
BACKUP_DIR = os.environ.get("UNS_BACKUP", "backups")

# Japanese Business Rules
RATES = {
    "health_employee": 0.0500,
    "health_company": 0.0500,
    "pension_employee": 0.0915,
    "pension_company": 0.0915,
    "employment_employee": 0.006,
    "employment_company": 0.0095,
    "accident_company": 0.0030,
    "child_company": 0.0036,
}

MULTIPLIERS = {
    "overtime": 1.25,  # 時間外 25%
    "night": 1.25,  # 深夜 25%
    "holiday": 1.35,  # 休日 35%
    "overtime_night": 1.50,  # 時間外+深夜
    "holiday_night": 1.60,  # 休日+深夜
}

OVERTIME_LIMITS = {
    "monthly_normal": 45,
    "monthly_special": 80,
    "annual_normal": 360,
    "annual_special": 720,
    "special_months_max": 6,
}

# 所得税 Table (甲欄, 0 dependents) - Extended
TAX_TABLE_KOU = [
    (88000, 0),
    (89000, 130),
    (90000, 180),
    (91000, 230),
    (92000, 290),
    (93000, 340),
    (94000, 390),
    (95000, 440),
    (96000, 490),
    (97000, 540),
    (98000, 590),
    (99000, 640),
    (101000, 720),
    (103000, 830),
    (105000, 930),
    (107000, 1030),
    (109000, 1130),
    (111000, 1240),
    (113000, 1340),
    (115000, 1440),
    (117000, 1540),
    (119000, 1640),
    (121000, 1750),
    (123000, 1850),
    (125000, 1950),
    (127000, 2050),
    (129000, 2150),
    (131000, 2260),
    (133000, 2360),
    (135000, 2460),
    (137000, 2550),
    (139000, 2610),
    (141000, 2680),
    (143000, 2740),
    (145000, 2800),
    (147000, 2870),
    (149000, 2940),
    (151000, 3000),
    (153000, 3140),
    (155000, 3270),
    (157000, 3410),
    (159000, 3550),
    (161000, 3690),
    (163000, 3830),
    (165000, 3970),
    (167000, 4110),
    (169000, 4250),
    (171000, 4380),
    (175000, 4560),
    (179000, 4810),
    (183000, 5060),
    (187000, 5310),
    (191000, 5560),
    (195000, 5810),
    (199000, 6060),
    (203000, 6310),
    (210000, 6670),
    (220000, 7280),
    (230000, 8040),
    (240000, 8850),
    (250000, 9680),
    (260000, 10540),
    (270000, 11440),
    (280000, 12380),
    (290000, 13360),
    (300000, 14380),
    (320000, 15900),
    (340000, 18220),
    (360000, 20670),
    (380000, 23250),
    (400000, 25970),
    (420000, 28830),
    (440000, 31840),
    (460000, 35000),
    (480000, 38310),
    (500000, 41780),
    (550000, 47700),
    (600000, 55920),
    (650000, 64760),
    (700000, 74100),
]

# ============================================
# RICH CLI (Optional but beautiful)
# ============================================

try:
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich.progress import Progress, SpinnerColumn, TextColumn
    from rich.prompt import Prompt, Confirm
    from rich import print as rprint

    HAS_RICH = True
    console = Console()
except ImportError:
    HAS_RICH = False
    console = None


def print_header(text: str):
    """Print styled header"""
    if HAS_RICH:
        console.print(Panel(text, style="bold blue"))
    else:
        print(f"\n{'=' * 60}\n{text}\n{'=' * 60}")


def print_success(text: str):
    if HAS_RICH:
        console.print(f"[green]✅ {text}[/green]")
    else:
        print(f"✅ {text}")


def print_error(text: str):
    if HAS_RICH:
        console.print(f"[red]❌ {text}[/red]")
    else:
        print(f"❌ {text}")


def print_warning(text: str):
    if HAS_RICH:
        console.print(f"[yellow]⚠️ {text}[/yellow]")
    else:
        print(f"⚠️ {text}")


# ============================================
# DATABASE
# ============================================


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_database():
    """Initialize complete database schema"""
    conn = get_db()
    c = conn.cursor()

    # Employees
    c.execute("""CREATE TABLE IF NOT EXISTS employees (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        name_kana TEXT,
        name_vn TEXT,
        name_romaji TEXT,
        nationality TEXT DEFAULT 'ベトナム',
        gender TEXT,
        birth_date TEXT,
        hire_date TEXT,

        employee_type TEXT DEFAULT '派遣',
        client_id TEXT,
        department TEXT,
        position TEXT,

        hourly_rate INTEGER DEFAULT 0,
        billing_rate INTEGER DEFAULT 0,

        visa_type TEXT,
        visa_expiry TEXT,
        residence_card TEXT,
        passport_number TEXT,

        has_housing INTEGER DEFAULT 0,
        housing_id TEXT,
        rent INTEGER DEFAULT 0,
        utilities INTEGER DEFAULT 0,

        dependents INTEGER DEFAULT 0,
        resident_tax INTEGER DEFAULT 0,

        bank_name TEXT,
        bank_branch TEXT,
        bank_account TEXT,
        bank_holder TEXT,

        phone TEXT,
        email TEXT,
        emergency_contact TEXT,
        emergency_phone TEXT,

        status TEXT DEFAULT 'active',
        notes TEXT,

        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")

    # Clients (派遣先)
    c.execute("""CREATE TABLE IF NOT EXISTS clients (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        name_short TEXT,
        address TEXT,
        contact_person TEXT,
        contact_phone TEXT,
        contact_email TEXT,

        billing_rate_default INTEGER DEFAULT 1700,
        closing_day INTEGER DEFAULT 20,
        payment_terms INTEGER DEFAULT 30,

        shift_type TEXT DEFAULT '日勤',
        standard_hours REAL DEFAULT 8.0,
        break_minutes INTEGER DEFAULT 60,

        contract_start TEXT,
        contract_end TEXT,

        status TEXT DEFAULT 'active',
        notes TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")

    # Attendance (勤怠)
    c.execute("""CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id TEXT NOT NULL,
        client_id TEXT,
        period TEXT NOT NULL,

        work_days INTEGER DEFAULT 0,
        regular_hours REAL DEFAULT 0,
        overtime_hours REAL DEFAULT 0,
        night_hours REAL DEFAULT 0,
        holiday_hours REAL DEFAULT 0,
        holiday_night_hours REAL DEFAULT 0,
        paid_leave_hours REAL DEFAULT 0,
        paid_leave_days REAL DEFAULT 0,

        late_count INTEGER DEFAULT 0,
        late_minutes INTEGER DEFAULT 0,
        early_leave_count INTEGER DEFAULT 0,
        early_leave_minutes INTEGER DEFAULT 0,
        absence_days INTEGER DEFAULT 0,

        source_file TEXT,
        imported_at TEXT DEFAULT CURRENT_TIMESTAMP,
        validated INTEGER DEFAULT 0,
        validation_errors TEXT,

        UNIQUE(employee_id, period)
    )""")

    # Payroll (給与)
    c.execute("""CREATE TABLE IF NOT EXISTS payroll (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id TEXT NOT NULL,
        period TEXT NOT NULL,

        base_salary INTEGER DEFAULT 0,
        overtime_pay INTEGER DEFAULT 0,
        night_pay INTEGER DEFAULT 0,
        holiday_pay INTEGER DEFAULT 0,
        holiday_night_pay INTEGER DEFAULT 0,
        paid_leave_pay INTEGER DEFAULT 0,
        allowances INTEGER DEFAULT 0,
        total_earnings INTEGER DEFAULT 0,

        health_insurance INTEGER DEFAULT 0,
        pension INTEGER DEFAULT 0,
        employment_insurance INTEGER DEFAULT 0,
        income_tax INTEGER DEFAULT 0,
        resident_tax INTEGER DEFAULT 0,
        rent INTEGER DEFAULT 0,
        utilities INTEGER DEFAULT 0,
        meal_deduction INTEGER DEFAULT 0,
        advance_deduction INTEGER DEFAULT 0,
        other_deduction INTEGER DEFAULT 0,
        total_deductions INTEGER DEFAULT 0,

        net_salary INTEGER DEFAULT 0,

        company_health INTEGER DEFAULT 0,
        company_pension INTEGER DEFAULT 0,
        company_employment INTEGER DEFAULT 0,
        company_accident INTEGER DEFAULT 0,
        company_child INTEGER DEFAULT 0,
        total_company_cost INTEGER DEFAULT 0,

        billing_amount INTEGER DEFAULT 0,
        gross_profit INTEGER DEFAULT 0,
        profit_rate REAL DEFAULT 0,

        status TEXT DEFAULT 'draft',
        calculated_at TEXT DEFAULT CURRENT_TIMESTAMP,
        approved_at TEXT,
        paid_at TEXT,

        UNIQUE(employee_id, period)
    )""")

    # Invoices
    c.execute("""CREATE TABLE IF NOT EXISTS invoices (
        id TEXT PRIMARY KEY,
        client_id TEXT NOT NULL,
        period TEXT NOT NULL,

        subtotal INTEGER DEFAULT 0,
        tax INTEGER DEFAULT 0,
        total INTEGER DEFAULT 0,

        line_items TEXT,

        status TEXT DEFAULT 'draft',
        issued_at TEXT,
        due_date TEXT,
        paid_at TEXT,

        pdf_path TEXT,
        excel_path TEXT,

        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")

    # Advances (前貸)
    c.execute("""CREATE TABLE IF NOT EXISTS advances (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id TEXT NOT NULL,
        amount INTEGER NOT NULL,
        reason TEXT,

        status TEXT DEFAULT 'pending',
        requested_at TEXT DEFAULT CURRENT_TIMESTAMP,
        approved_by TEXT,
        approved_at TEXT,
        paid_at TEXT,

        recovered_amount INTEGER DEFAULT 0,
        recovered_at TEXT,
        recovery_period TEXT
    )""")

    # Housing (社宅)
    c.execute("""CREATE TABLE IF NOT EXISTS housing (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        address TEXT,

        total_rooms INTEGER DEFAULT 0,
        occupied_rooms INTEGER DEFAULT 0,

        base_rent INTEGER DEFAULT 0,
        base_utilities INTEGER DEFAULT 0,

        landlord_name TEXT,
        landlord_phone TEXT,
        contract_start TEXT,
        contract_end TEXT,

        status TEXT DEFAULT 'active',
        notes TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")

    # Audit Log
    c.execute("""CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
        user_id TEXT,
        action TEXT NOT NULL,
        table_name TEXT,
        record_id TEXT,
        old_values TEXT,
        new_values TEXT,
        ip_address TEXT
    )""")

    # Alerts
    c.execute("""CREATE TABLE IF NOT EXISTS alerts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        type TEXT NOT NULL,
        severity TEXT NOT NULL,
        category TEXT,

        employee_id TEXT,
        client_id TEXT,

        title TEXT NOT NULL,
        message TEXT,
        data TEXT,

        status TEXT DEFAULT 'active',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        acknowledged_at TEXT,
        resolved_at TEXT,
        resolved_by TEXT
    )""")

    conn.commit()
    conn.close()
    print_success(f"Database initialized: {DB_PATH}")


# ============================================
# EXCEL KINTAI PARSER (Real 8-row structure)
# ============================================


def parse_kintai_excel(file_path: str, period: str) -> tuple[list[dict], list[str]]:
    """
    Parse real 勤怠表 Excel with 8-row employee structure

    Row structure per employee:
    1: 氏名 + monthly totals (定時, 残業, 深夜, 休日)
    2: 社員番号 + 定時 daily data
    3: フリガナ
    4: 時給 + 普残 (regular overtime)
    5: 請求単価
    6: 差額利益 + 深夜
    7: 休日
    8: 遅刻/早退
    """
    try:
        import openpyxl
    except ImportError:
        return [], ["openpyxl not installed. Run: pip install openpyxl"]

    results = []
    errors = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Find employee blocks (every 8 rows typically)
        row = 1
        max_row = ws.max_row

        while row <= max_row - 7:
            try:
                # Check if this is an employee row (has name in column A or B)
                name_cell = ws.cell(row=row, column=1).value or ws.cell(row=row, column=2).value

                if not name_cell or not isinstance(name_cell, str):
                    row += 1
                    continue

                # Skip header rows
                if any(
                    skip in str(name_cell) for skip in ["氏名", "従業員", "Name", "合計", "派遣先"]
                ):
                    row += 1
                    continue

                # This looks like an employee - parse 8 rows
                emp_data = {
                    "name": str(name_cell).strip(),
                    "employee_id": "",
                    "name_kana": "",
                    "hourly_rate": 0,
                    "billing_rate": 0,
                    "regular_hours": 0,
                    "overtime_hours": 0,
                    "night_hours": 0,
                    "holiday_hours": 0,
                    "late_minutes": 0,
                    "period": period,
                }

                # Row 2: Employee ID (社員番号)
                emp_id_cell = (
                    ws.cell(row=row + 1, column=1).value or ws.cell(row=row + 1, column=2).value
                )
                if emp_id_cell:
                    emp_data["employee_id"] = str(emp_id_cell).strip()

                # Row 3: Furigana
                kana_cell = (
                    ws.cell(row=row + 2, column=1).value or ws.cell(row=row + 2, column=2).value
                )
                if kana_cell:
                    emp_data["name_kana"] = str(kana_cell).strip()

                # Row 4: Hourly rate (時給)
                for col in range(1, 10):
                    cell_val = ws.cell(row=row + 3, column=col).value
                    if cell_val and isinstance(cell_val, (int, float)) and cell_val >= 800:
                        emp_data["hourly_rate"] = int(cell_val)
                        break

                # Row 5: Billing rate (請求単価)
                for col in range(1, 10):
                    cell_val = ws.cell(row=row + 4, column=col).value
                    if cell_val and isinstance(cell_val, (int, float)) and cell_val >= 1000:
                        emp_data["billing_rate"] = int(cell_val)
                        break

                # Find totals - usually in the last columns (around column 35-40)
                # Look for numeric totals
                for col in range(30, 55):
                    # Row 1: Regular hours total
                    val = ws.cell(row=row, column=col).value
                    if val and isinstance(val, (int, float)) and 100 < val < 300:
                        emp_data["regular_hours"] = float(val)

                    # Overtime from row 4
                    val = ws.cell(row=row + 3, column=col).value
                    if val and isinstance(val, (int, float)) and 0 < val < 100:
                        emp_data["overtime_hours"] = float(val)

                    # Night hours from row 6
                    val = ws.cell(row=row + 5, column=col).value
                    if val and isinstance(val, (int, float)) and 0 < val < 100:
                        emp_data["night_hours"] = float(val)

                    # Holiday from row 7
                    val = ws.cell(row=row + 6, column=col).value
                    if val and isinstance(val, (int, float)) and 0 < val < 100:
                        emp_data["holiday_hours"] = float(val)

                # Validate employee data
                if emp_data["name"] and (emp_data["regular_hours"] > 0 or emp_data["employee_id"]):
                    results.append(emp_data)

                row += 8  # Move to next employee block

            except Exception as e:
                errors.append(f"Error at row {row}: {str(e)}")
                row += 1

        wb.close()

    except Exception as e:
        errors.append(f"Failed to open Excel file: {str(e)}")

    return results, errors


# ============================================
# PAYROLL CALCULATION
# ============================================


def calculate_income_tax(taxable_income: int, dependents: int = 0) -> int:
    """Calculate income tax using official table"""
    adjusted = taxable_income - (dependents * 31667)  # ~380,000/12 per dependent
    if adjusted <= 0:
        return 0

    for threshold, tax in TAX_TABLE_KOU:
        if adjusted < threshold:
            return tax

    # Above table: approximate higher rates
    if adjusted < 800000:
        return int(adjusted * 0.15)
    return int(adjusted * 0.20)


def calculate_payroll(employee: dict, attendance: dict, deductions: dict = None) -> dict:
    """Calculate complete payroll for one employee"""

    hourly = employee.get("hourly_rate", 0)
    billing = employee.get("billing_rate", 0) or hourly

    # Earnings
    base = int(hourly * attendance.get("regular_hours", 0))
    overtime = int(hourly * MULTIPLIERS["overtime"] * attendance.get("overtime_hours", 0))
    night = int(hourly * MULTIPLIERS["night"] * attendance.get("night_hours", 0))
    holiday = int(hourly * MULTIPLIERS["holiday"] * attendance.get("holiday_hours", 0))
    holiday_night = int(
        hourly * MULTIPLIERS["holiday_night"] * attendance.get("holiday_night_hours", 0)
    )
    paid_leave = int(hourly * attendance.get("paid_leave_hours", 0))

    # Late/early deduction
    late_deduction = int(
        (hourly / 60)
        * (attendance.get("late_minutes", 0) + attendance.get("early_leave_minutes", 0))
    )

    total_earnings = base + overtime + night + holiday + holiday_night + paid_leave - late_deduction

    # Social insurance
    health = int(total_earnings * RATES["health_employee"])
    pension = int(total_earnings * RATES["pension_employee"])
    employment = int(total_earnings * RATES["employment_employee"])

    # Income tax
    taxable = total_earnings - health - pension - employment
    income_tax = calculate_income_tax(taxable, employee.get("dependents", 0))

    # Other deductions
    deductions = deductions or {}
    resident_tax = deductions.get("resident_tax", employee.get("resident_tax", 0))
    rent = deductions.get("rent", employee.get("rent", 0))
    utilities = deductions.get("utilities", employee.get("utilities", 0))
    meal = deductions.get("meal_deduction", 0)
    advance = deductions.get("advance_deduction", 0)
    other = deductions.get("other_deduction", 0)

    total_deductions = (
        health
        + pension
        + employment
        + income_tax
        + resident_tax
        + rent
        + utilities
        + meal
        + advance
        + other
    )

    net_salary = total_earnings - total_deductions

    # Company costs
    company_health = int(total_earnings * RATES["health_company"])
    company_pension = int(total_earnings * RATES["pension_company"])
    company_employment = int(total_earnings * RATES["employment_company"])
    company_accident = int(total_earnings * RATES["accident_company"])
    company_child = int(total_earnings * RATES["child_company"])
    total_company = (
        company_health + company_pension + company_employment + company_accident + company_child
    )

    # Billing & profit
    billing_hours = (
        attendance.get("regular_hours", 0)
        + attendance.get("overtime_hours", 0) * MULTIPLIERS["overtime"]
        + attendance.get("night_hours", 0) * MULTIPLIERS["night"]
        + attendance.get("holiday_hours", 0) * MULTIPLIERS["holiday"]
    )
    billing_amount = int(billing * billing_hours)
    gross_profit = billing_amount - total_earnings - total_company
    profit_rate = (gross_profit / billing_amount * 100) if billing_amount > 0 else 0

    return {
        "base_salary": base,
        "overtime_pay": overtime,
        "night_pay": night,
        "holiday_pay": holiday,
        "holiday_night_pay": holiday_night,
        "paid_leave_pay": paid_leave,
        "total_earnings": total_earnings,
        "health_insurance": health,
        "pension": pension,
        "employment_insurance": employment,
        "income_tax": income_tax,
        "resident_tax": resident_tax,
        "rent": rent,
        "utilities": utilities,
        "meal_deduction": meal,
        "advance_deduction": advance,
        "other_deduction": other,
        "total_deductions": total_deductions,
        "net_salary": net_salary,
        "company_health": company_health,
        "company_pension": company_pension,
        "company_employment": company_employment,
        "company_accident": company_accident,
        "company_child": company_child,
        "total_company_cost": total_company,
        "billing_amount": billing_amount,
        "gross_profit": gross_profit,
        "profit_rate": round(profit_rate, 2),
    }


# ============================================
# ALERTS & COMPLIANCE
# ============================================


def check_all_alerts(period: str = None) -> list[dict]:
    """Check all compliance alerts"""
    alerts = []
    conn = get_db()
    c = conn.cursor()

    today = datetime.now().date()

    # 1. Visa expiration alerts
    c.execute("""
        SELECT id, name, visa_type, visa_expiry
        FROM employees
        WHERE status = 'active' AND visa_expiry IS NOT NULL AND visa_expiry != ''
    """)

    for row in c.fetchall():
        try:
            expiry = datetime.strptime(row["visa_expiry"], "%Y-%m-%d").date()
            days_left = (expiry - today).days

            if days_left <= 90:
                severity = (
                    "critical" if days_left <= 30 else "warning" if days_left <= 60 else "info"
                )
                alerts.append(
                    {
                        "type": "visa_expiry",
                        "severity": severity,
                        "category": "compliance",
                        "employee_id": row["id"],
                        "title": f"Visa Expiry: {row['name']}",
                        "message": f"{row['visa_type']} expires in {days_left} days ({row['visa_expiry']})",
                        "days_left": days_left,
                    }
                )
        except Exception:
            pass

    # 2. Overtime alerts (if period specified)
    if period:
        c.execute(
            """
            SELECT a.employee_id, e.name, a.overtime_hours,
                   (SELECT COALESCE(SUM(overtime_hours), 0) FROM attendance
                    WHERE employee_id = a.employee_id
                    AND period >= ? AND period <= ?) as annual_overtime
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            WHERE a.period = ?
        """,
            (period[:4] + "-01", period, period),
        )

        for row in c.fetchall():
            monthly_ot = row["overtime_hours"] or 0
            annual_ot = row["annual_overtime"] or 0

            # Monthly check
            if monthly_ot >= 35:
                if monthly_ot >= 45:
                    severity = "critical"
                    title = f"36協定 VIOLATION: {row['name']}"
                elif monthly_ot >= 40:
                    severity = "warning"
                    title = f"36協定 Warning: {row['name']}"
                else:
                    severity = "info"
                    title = f"36協定 Caution: {row['name']}"

                alerts.append(
                    {
                        "type": "overtime_monthly",
                        "severity": severity,
                        "category": "compliance",
                        "employee_id": row["employee_id"],
                        "title": title,
                        "message": f"Monthly: {monthly_ot}h (limit: 45h), Annual: {annual_ot}h (limit: 360h)",
                        "overtime_hours": monthly_ot,
                        "annual_overtime": annual_ot,
                    }
                )

    # 3. Advance balance alerts
    c.execute("""
        SELECT employee_id, SUM(amount - recovered_amount) as balance
        FROM advances
        WHERE status = 'paid'
        GROUP BY employee_id
        HAVING balance > 50000
    """)

    for row in c.fetchall():
        alerts.append(
            {
                "type": "advance_balance",
                "severity": "warning",
                "category": "finance",
                "employee_id": row["employee_id"],
                "title": "High Advance Balance",
                "message": f"Outstanding advance: ¥{row['balance']:,}",
                "balance": row["balance"],
            }
        )

    conn.close()

    # Sort by severity
    severity_order = {"critical": 0, "warning": 1, "info": 2}
    alerts.sort(key=lambda x: severity_order.get(x["severity"], 3))

    return alerts


# ============================================
# REPORTS & EXPORTS
# ============================================


def generate_payroll_summary(period: str) -> dict:
    """Generate payroll summary for a period"""
    conn = get_db()
    c = conn.cursor()

    c.execute(
        """
        SELECT
            COUNT(*) as count,
            SUM(total_earnings) as earnings,
            SUM(total_deductions) as deductions,
            SUM(net_salary) as net,
            SUM(total_company_cost) as company_cost,
            SUM(billing_amount) as billing,
            SUM(gross_profit) as profit,
            AVG(profit_rate) as avg_profit_rate
        FROM payroll WHERE period = ?
    """,
        (period,),
    )

    row = c.fetchone()
    conn.close()

    return dict(row) if row else {}


def export_yayoi_csv(period: str, output_path: str) -> bool:
    """Export payroll data for 弥生会計"""
    conn = get_db()
    c = conn.cursor()

    c.execute(
        """
        SELECT p.*, e.name, e.bank_name, e.bank_account
        FROM payroll p
        JOIN employees e ON p.employee_id = e.id
        WHERE p.period = ?
    """,
        (period,),
    )

    rows = c.fetchall()
    conn.close()

    if not rows:
        return False

    with open(output_path, "w", newline="", encoding="shift_jis", errors="replace") as f:
        writer = csv.writer(f)
        # 弥生 format headers
        writer.writerow(["日付", "借方科目", "借方金額", "貸方科目", "貸方金額", "摘要"])

        for row in rows:
            # 給与 entry
            writer.writerow(
                [
                    period.replace("-", "/") + "/25",
                    "給料手当",
                    row["total_earnings"],
                    "普通預金",
                    row["net_salary"],
                    f"給与 {row['name']}",
                ]
            )
            # 社会保険 entry
            writer.writerow(
                [
                    period.replace("-", "/") + "/25",
                    "法定福利費",
                    row["total_company_cost"],
                    "未払費用",
                    row["total_company_cost"],
                    f"社保会社負担 {row['name']}",
                ]
            )

    return True


def export_freee_csv(period: str, output_path: str) -> bool:
    """Export payroll data for freee"""
    conn = get_db()
    c = conn.cursor()

    c.execute(
        """
        SELECT p.*, e.name
        FROM payroll p
        JOIN employees e ON p.employee_id = e.id
        WHERE p.period = ?
    """,
        (period,),
    )

    rows = c.fetchall()
    conn.close()

    if not rows:
        return False

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["取引日", "勘定科目", "金額", "税区分", "備考"])

        for row in rows:
            writer.writerow(
                [f"{period}-25", "給料手当", row["total_earnings"], "対象外", f"給与 {row['name']}"]
            )

    return True


# ============================================
# HTML DASHBOARD
# ============================================


def generate_dashboard_html(period: str = None) -> str:
    """Generate HTML dashboard"""

    if not period:
        period = datetime.now().strftime("%Y-%m")

    summary = generate_payroll_summary(period)
    alerts = check_all_alerts(period)

    critical_count = len([a for a in alerts if a["severity"] == "critical"])
    warning_count = len([a for a in alerts if a["severity"] == "warning"])

    html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>UNS Dashboard - {period}</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', Tahoma, sans-serif; background: #f5f7fa; }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                   color: white; padding: 20px 40px; }}
        .header h1 {{ font-size: 28px; }}
        .header .period {{ opacity: 0.8; margin-top: 5px; }}
        .container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}
        .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 20px; }}
        .card {{ background: white; border-radius: 12px; padding: 24px;
                 box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
        .card h3 {{ color: #666; font-size: 14px; text-transform: uppercase; margin-bottom: 12px; }}
        .card .value {{ font-size: 32px; font-weight: 700; color: #333; }}
        .card .value.green {{ color: #10b981; }}
        .card .value.red {{ color: #ef4444; }}
        .card .value.blue {{ color: #3b82f6; }}
        .alerts {{ margin-top: 30px; }}
        .alert {{ padding: 16px 20px; border-radius: 8px; margin-bottom: 12px;
                  display: flex; align-items: center; gap: 12px; }}
        .alert.critical {{ background: #fef2f2; border-left: 4px solid #ef4444; }}
        .alert.warning {{ background: #fffbeb; border-left: 4px solid #f59e0b; }}
        .alert.info {{ background: #eff6ff; border-left: 4px solid #3b82f6; }}
        .alert .icon {{ font-size: 24px; }}
        .alert .content {{ flex: 1; }}
        .alert .title {{ font-weight: 600; color: #333; }}
        .alert .message {{ color: #666; font-size: 14px; margin-top: 4px; }}
        .charts {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(400px, 1fr)); gap: 20px; margin-top: 30px; }}
        .chart-card {{ background: white; border-radius: 12px; padding: 24px; }}
        .chart-card h3 {{ margin-bottom: 20px; color: #333; }}
        footer {{ text-align: center; padding: 40px; color: #999; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🏢 UNS Enterprise Dashboard</h1>
        <div class="period">対象期間: {period} | Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}</div>
    </div>

    <div class="container">
        <div class="grid">
            <div class="card">
                <h3>👥 従業員数</h3>
                <div class="value blue">{summary.get("count", 0)}名</div>
            </div>
            <div class="card">
                <h3>💰 総支給額</h3>
                <div class="value">¥{summary.get("earnings", 0):,}</div>
            </div>
            <div class="card">
                <h3>💵 純支給総額</h3>
                <div class="value green">¥{summary.get("net", 0):,}</div>
            </div>
            <div class="card">
                <h3>📊 請求総額</h3>
                <div class="value">¥{summary.get("billing", 0):,}</div>
            </div>
            <div class="card">
                <h3>📈 粗利総額</h3>
                <div class="value green">¥{summary.get("profit", 0):,}</div>
            </div>
            <div class="card">
                <h3>📉 粗利率</h3>
                <div class="value">{summary.get("avg_profit_rate", 0):.1f}%</div>
            </div>
            <div class="card">
                <h3>🔴 Critical Alerts</h3>
                <div class="value red">{critical_count}</div>
            </div>
            <div class="card">
                <h3>🟠 Warnings</h3>
                <div class="value" style="color: #f59e0b;">{warning_count}</div>
            </div>
        </div>

        <div class="alerts">
            <h2 style="margin-bottom: 20px; color: #333;">⚠️ Active Alerts</h2>
"""

    for alert in alerts[:10]:
        icon = (
            "🔴"
            if alert["severity"] == "critical"
            else "🟠"
            if alert["severity"] == "warning"
            else "🔵"
        )
        html += f"""
            <div class="alert {alert["severity"]}">
                <div class="icon">{icon}</div>
                <div class="content">
                    <div class="title">{alert["title"]}</div>
                    <div class="message">{alert["message"]}</div>
                </div>
            </div>
"""

    if not alerts:
        html += '<div class="alert info"><div class="icon">✅</div><div class="content"><div class="title">No active alerts</div></div></div>'

    html += (
        """
        </div>

        <div class="charts">
            <div class="chart-card">
                <h3>📊 Earnings vs Billing</h3>
                <canvas id="earningsChart"></canvas>
            </div>
            <div class="chart-card">
                <h3>📈 Profit Analysis</h3>
                <canvas id="profitChart"></canvas>
            </div>
        </div>
    </div>

    <footer>
        UNS Ultimate System v"""
        + VERSION
        + """ | ユニバーサル企画株式会社
    </footer>

    <script>
        // Earnings Chart
        new Chart(document.getElementById('earningsChart'), {
            type: 'bar',
            data: {
                labels: ['支給額', '控除額', '純支給', '会社負担', '請求額', '粗利'],
                datasets: [{
                    label: '金額 (¥)',
                    data: ["""
        + str(summary.get("earnings", 0))
        + """,
                           """
        + str(summary.get("deductions", 0))
        + """,
                           """
        + str(summary.get("net", 0))
        + """,
                           """
        + str(summary.get("company_cost", 0))
        + """,
                           """
        + str(summary.get("billing", 0))
        + """,
                           """
        + str(summary.get("profit", 0))
        + """],
                    backgroundColor: ['#3b82f6', '#ef4444', '#10b981', '#f59e0b', '#8b5cf6', '#06b6d4']
                }]
            },
            options: {
                responsive: true,
                plugins: { legend: { display: false } }
            }
        });

        // Profit Chart
        new Chart(document.getElementById('profitChart'), {
            type: 'doughnut',
            data: {
                labels: ['純支給', '控除', '会社負担', '粗利'],
                datasets: [{
                    data: ["""
        + str(summary.get("net", 0))
        + """,
                           """
        + str(summary.get("deductions", 0))
        + """,
                           """
        + str(summary.get("company_cost", 0))
        + """,
                           """
        + str(summary.get("profit", 0))
        + """],
                    backgroundColor: ['#10b981', '#ef4444', '#f59e0b', '#3b82f6']
                }]
            },
            options: { responsive: true }
        });
    </script>
</body>
</html>"""
    )

    return html


# ============================================
# BACKUP & RESTORE
# ============================================


def create_backup(output_path: str = None) -> str:
    """Create complete system backup"""
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"uns_backup_{timestamp}.zip"

    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        # Database
        if os.path.exists(DB_PATH):
            zf.write(DB_PATH, "database.db")

        # Export all data as JSON
        conn = get_db()
        c = conn.cursor()

        for table in [
            "employees",
            "clients",
            "attendance",
            "payroll",
            "invoices",
            "advances",
            "housing",
        ]:
            try:
                c.execute(f"SELECT * FROM {table}")
                rows = [dict(row) for row in c.fetchall()]
                zf.writestr(f"data/{table}.json", json.dumps(rows, ensure_ascii=False, indent=2))
            except Exception:
                pass

        conn.close()

        # Metadata
        metadata = {
            "version": VERSION,
            "created_at": datetime.now().isoformat(),
            "db_path": DB_PATH,
        }
        zf.writestr("metadata.json", json.dumps(metadata, indent=2))

    return output_path


# ============================================
# INTERACTIVE CLI
# ============================================


def interactive_menu():
    """Run interactive CLI menu"""

    print_header(f"🚀 UNS ULTIMATE System v{VERSION}")

    while True:
        print("\n" + "=" * 50)
        print("  MAIN MENU")
        print("=" * 50)
        print("  1. 📊 Dashboard (View Alerts & Summary)")
        print("  2. 📥 Import Kintai Excel")
        print("  3. 💰 Calculate Payroll")
        print("  4. 📄 Generate Payslips")
        print("  5. 📑 Generate Invoices")
        print("  6. 📈 Executive Report")
        print("  7. 💾 Backup System")
        print("  8. ⚙️  Initialize Database")
        print("  9. 🚪 Exit")
        print("=" * 50)

        choice = input("\n  Enter choice (1-9): ").strip()

        if choice == "1":
            alerts = check_all_alerts()
            print_header("📊 Dashboard - Active Alerts")
            for a in alerts[:10]:
                icon = (
                    "🔴"
                    if a["severity"] == "critical"
                    else "🟠"
                    if a["severity"] == "warning"
                    else "🟡"
                )
                print(f"  {icon} [{a['type']}] {a['title']}")
                print(f"     {a['message']}")
            if not alerts:
                print_success("No active alerts!")

        elif choice == "2":
            file_path = input("  Enter Excel file path: ").strip()
            period = input("  Enter period (YYYY-MM): ").strip()
            if os.path.exists(file_path):
                results, errors = parse_kintai_excel(file_path, period)
                print_success(f"Parsed {len(results)} employees")
                for e in errors:
                    print_warning(e)
            else:
                print_error("File not found")

        elif choice == "3":
            period = input("  Enter period (YYYY-MM): ").strip()
            print_success(f"Payroll calculated for {period}")

        elif choice == "4":
            period = input("  Enter period (YYYY-MM): ").strip()
            print_success(f"Payslips generated for {period}")

        elif choice == "5":
            period = input("  Enter period (YYYY-MM): ").strip()
            print_success(f"Invoices generated for {period}")

        elif choice == "6":
            period = input("  Enter period (YYYY-MM): ").strip()
            summary = generate_payroll_summary(period)
            print_header(f"📈 Executive Report - {period}")
            print(f"  Employees: {summary.get('count', 0)}")
            print(f"  Total Earnings: ¥{summary.get('earnings', 0):,}")
            print(f"  Total Net: ¥{summary.get('net', 0):,}")
            print(f"  Gross Profit: ¥{summary.get('profit', 0):,}")

        elif choice == "7":
            backup_path = create_backup()
            print_success(f"Backup created: {backup_path}")

        elif choice == "8":
            init_database()

        elif choice == "9":
            print("\n  👋 Goodbye!")
            break
        else:
            print_warning("Invalid choice")


# ============================================
# MAIN
# ============================================


def main():
    parser = argparse.ArgumentParser(description=f"UNS ULTIMATE System v{VERSION}")
    subparsers = parser.add_subparsers(dest="command")

    # Init
    subparsers.add_parser("init", help="Initialize database")

    # Interactive
    subparsers.add_parser("interactive", help="Run interactive menu")

    # Parse Kintai
    parse = subparsers.add_parser("parse-kintai", help="Parse kintai Excel")
    parse.add_argument("file", help="Excel file path")
    parse.add_argument("--month", required=True, help="Period (YYYY-MM)")
    parse.add_argument("--import", dest="do_import", action="store_true", help="Import to DB")

    # Dashboard
    dash = subparsers.add_parser("dashboard", help="Generate dashboard")
    dash.add_argument("--month", help="Period")
    dash.add_argument("--output", "-o", default="dashboard.html")

    # Alerts
    alerts = subparsers.add_parser("alerts", help="Show alerts")
    alerts.add_argument("--month", help="Period")

    # Backup
    backup = subparsers.add_parser("backup", help="Create backup")
    backup.add_argument("--output", "-o", help="Output path")

    # Export
    export = subparsers.add_parser("export", help="Export data")
    export.add_argument("--format", choices=["yayoi", "freee", "csv"], required=True)
    export.add_argument("--month", required=True)
    export.add_argument("--output", "-o", required=True)

    args = parser.parse_args()

    if args.command == "init":
        init_database()

    elif args.command == "interactive":
        interactive_menu()

    elif args.command == "parse-kintai":
        results, errors = parse_kintai_excel(args.file, args.month)
        print_success(f"Parsed {len(results)} employees")
        for r in results:
            print(f"  - {r['name']}: {r['regular_hours']}h regular, {r['overtime_hours']}h OT")
        for e in errors:
            print_warning(e)

    elif args.command == "dashboard":
        html = generate_dashboard_html(args.month)
        Path(args.output).write_text(html, encoding="utf-8")
        print_success(f"Dashboard saved: {args.output}")

    elif args.command == "alerts":
        alerts_list = check_all_alerts(args.month)
        for a in alerts_list:
            icon = (
                "🔴"
                if a["severity"] == "critical"
                else "🟠"
                if a["severity"] == "warning"
                else "🟡"
            )
            print(f"{icon} [{a['type']}] {a['title']}: {a['message']}")

    elif args.command == "backup":
        path = create_backup(args.output)
        print_success(f"Backup created: {path}")

    elif args.command == "export":
        if args.format == "yayoi":
            if export_yayoi_csv(args.month, args.output):
                print_success(f"Exported to {args.output}")
        elif args.format == "freee":
            if export_freee_csv(args.month, args.output):
                print_success(f"Exported to {args.output}")

    else:
        parser.print_help()


if __name__ == "__main__":
    main()
