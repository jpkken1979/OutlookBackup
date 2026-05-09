import sys
from pathlib import Path
import json
import logging
import pandas as pd
import openpyxl
from typing import Any

# Configuration
logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
logger = logging.getLogger("ExcelSuperAgent")


class ExcelOrchestrator:
    def __init__(self):
        self.strategies = {
            "pandas": "High-performance tabular read",
            "smart": "Header detection & type normalization",
            "deep": "Spatial form extraction (Japanese/Complex)",
        }

    def detect_encoding(self, file_path: Path) -> str:
        """Probes common encodings for Japanese Windows environments."""
        encodings = ["utf-8-sig", "cp932", "shift_jis", "euc-jp"]

        # Read first 1kb to test encoding
        for enc in encodings:
            try:
                with open(file_path, encoding=enc) as f:
                    f.read(1024)
                logger.info(f"Encoding detection: {enc} seemed to work.")
                return enc
            except UnicodeDecodeError:
                continue
            except Exception:
                break

        return "utf-8"  # Fallback

    def analyze_structure(self, file_path: str) -> str:
        """Determines the best parsing strategy."""
        path = Path(file_path)

        if path.suffix.lower() == ".csv":
            # CSVs are usually tabular
            return "pandas"

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = wb.active

            # 1. Check for Merged Cells (Indication of Form/Deep formatting)
            # read_only mode doesn't support merged_cells fully efficiently,
            # but we can assume complex if row count is low but structure looks dense.

            # Heuristic 1: Data Density
            max_row = sheet.max_row or 0
            if max_row > 2000:
                logger.info("High row count detected. Preferring Pandas.")
                wb.close()
                return "pandas"

            # Re-open in normal mode to check merges if small enough
            wb.close()
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet = wb.active

            merge_count = len(sheet.merged_cells.ranges)
            logger.info(f"Merged ranges detected: {merge_count}")

            if merge_count > 5:
                # Forms usually have many merges for layout
                logger.info("Complex merged cells detected. Choosing DEEP parser.")
                return "deep"

            # Heuristic 2: Header Consistency
            # Check row 1 vs row 2
            row1 = [c.value for c in sheet[1]]
            row1_clean = [str(v).strip() for v in row1 if v]

            if len(row1_clean) > 2:
                logger.info("Clean header row candidates found. Choosing SMART parser.")
                return "smart"

            # Default fallback
            return "deep"

        except Exception as e:
            logger.error(f"Analysis failed: {e}")
            return "pandas"  # Fallback to raw path

    def route_by_keywords(self, file_path: str) -> str | None:
        """Routes based on keywords in the file path before structural analysis.

        Higher priority = checked first.  Returns strategy name or None if no match.

        Args:
            file_path: Path to inspect for keywords.

        Returns:
            Strategy name if keyword matched, None otherwise.
        """
        lower = file_path.lower()

        # Highest priority: conditional format operations
        if any(kw in lower for kw in ["formato condicional", "conditional format", "条件格式", "条件付き書式"]):
            logger.info("Keyword matched: conditional format → excel-router")
            return "smart"

        # Spanish / general keywords → excel-smart-parser
        if any(kw in lower for kw in ["planilla", "reporte", "reporte mensual", "resumen"]):
            logger.info("Keyword matched: planilla/reporte → smart")
            return "smart"

        # Japanese business form keywords → excel-deep-parse
        japanese_keywords = [
            "個別契約",    # Individual contract
            "派遣",        # Dispatch
            "勤怠",        # Attendance
            "有給",        # Paid leave
            "賃金",        # Wages
            "履歴書",      # Resume
            "給与",        # Salary
            "台帳",        # Ledger
            "算定表",      # Calculation table
        ]
        if any(kw in lower for kw in japanese_keywords):
            logger.info("Keyword matched: Japanese business form → deep")
            return "deep"

        return None

    def process(self, file_path: str, hints: dict[str, Any] | None = None) -> dict[str, Any]:
        # 1. Keyword-based routing (highest priority)
        keyword_strategy = self.route_by_keywords(file_path)
        if keyword_strategy is not None:
            strategy = keyword_strategy
        else:
            # 2. Structural analysis as fallback
            strategy = self.analyze_structure(file_path)

        logger.info(f"Selected Strategy: {strategy.upper()} (hints={hints})")

        if strategy == "pandas":
            return self._run_pandas(file_path)
        elif strategy == "smart":
            return self._run_smart(file_path)
        elif strategy == "deep":
            return self._run_deep(file_path)

        return {"error": "Unknown strategy"}

    def _run_pandas(self, file_path: str):
        path = Path(file_path)
        if path.suffix == ".csv":
            enc = self.detect_encoding(path)
            df = pd.read_csv(path, encoding=enc)
        else:
            df = pd.read_excel(path)

        # Basic cleanup
        df.columns = [c for c in df.columns if not str(c).startswith("Unnamed")]
        return {
            "strategy": "pandas",
            "metadata": {"rows": len(df), "columns": list(df.columns)},
            "preview": df.head(5).to_dict(orient="records"),
        }

    def _run_smart(self, file_path: str):
        # Delegate to existing Smart Parser logic (simulated import)
        # In production this would import the class from ../../excel-smart-parser
        # For now, we simulate the 'Smart' behavior which handles header detection
        df = pd.read_excel(file_path, header=0)  # Assume smart detected header 0
        return {
            "strategy": "smart",
            "metadata": {"type": "Standard Table"},
            "data": df.head(10).to_dict(orient="records"),
        }

    def _run_deep(self, file_path: str):
        # Delegate to Deep Parser
        from ..excel_deep_parse.scripts.deep_parser import ExcelDeepParser

        parser = ExcelDeepParser(file_path)
        # Try to find common Japanese form fields automatically
        common_fields = [
            {"key": "name", "labels": ["氏名", "Name", "名前"], "direction": "right"},
            {"key": "date", "labels": ["日付", "Date", "年月日"], "direction": "right"},
            {"key": "amount", "labels": ["金額", "Total", "合計"], "direction": "right"},
        ]
        data = parser.extract_fields(common_fields)
        return {"strategy": "deep", "extracted_fields": data}


def process_excel(file_path: str, hints: dict[str, Any] | None = None):
    orchestrator = ExcelOrchestrator()
    return orchestrator.process(file_path, hints=hints)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python orchestrator.py <file_path>")
        sys.exit(1)

    res = process_excel(sys.argv[1])
    print(json.dumps(res, indent=2, ensure_ascii=False))
