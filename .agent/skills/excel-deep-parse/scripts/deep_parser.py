import openpyxl
import json
from dataclasses import dataclass
from typing import Any


@dataclass
class DeepCell:
    value: Any
    row: int
    col: int
    coordinate: str
    bg_color: str | None
    is_merged: bool
    merged_range: str | None
    font_bold: bool


class ExcelDeepParser:
    def __init__(self, file_path: str):
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.file_path = file_path

    def _get_cell_info(self, sheet, cell) -> DeepCell:
        # Handle RGB colors (openpyxl sometimes gives theme/tint)
        bg_color = None
        if cell.fill and cell.fill.start_color:
            if cell.fill.start_color.type == "rgb":
                bg_color = cell.fill.start_color.rgb

        is_merged = False
        merged_range = None
        for rng in sheet.merged_cells.ranges:
            if cell.coordinate in rng:
                is_merged = True
                merged_range = str(rng)
                break

        return DeepCell(
            value=cell.value,
            row=cell.row,
            col=cell.column,
            coordinate=cell.coordinate,
            bg_color=bg_color,
            is_merged=is_merged,
            merged_range=merged_range,
            font_bold=cell.font.bold if cell.font else False,
        )

    def extract_fields(
        self, field_specs: list[dict[str, Any]], sheet_name: str | None = None
    ) -> dict[str, Any]:
        """
        Extracts key-value pairs based on spatial proximity.

        field_specs example:
        [
            {"key": "employee_name", "labels": ["氏名", "Name"], "direction": "right"},
            {"key": "total_amount", "labels": ["Total"], "direction": "below"}
        ]
        """
        if sheet_name:
            sheet = self.wb[sheet_name]
        else:
            sheet = self.wb.active

        results = {}

        # Iterate efficiently? For forms, maybe just iterate used range.
        # Create a map of cell values to coordinates for label finding
        cell_map = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    txt = str(cell.value).strip().lower()
                    if txt not in cell_map:
                        cell_map[txt] = []
                    cell_map[txt].append(cell)

        for spec in field_specs:
            labels = spec.get("labels", [])
            direction = spec.get("direction", "auto")  # auto, right, below
            found_val = None

            for label in labels:
                clean_label = label.lower()
                candidates = []

                # Loose matching
                for txt, cells in cell_map.items():
                    if clean_label in txt:
                        candidates.extend(cells)

                if not candidates:
                    continue

                # Sort candidates by exact match length (shortest first => closest match)
                candidates.sort(key=lambda c: len(str(c.value)))

                target_label_cell = candidates[0]

                # Look for value based on direction
                val_cell = None

                if direction in ["right", "auto"]:
                    # Try right neighbor (col + 1)
                    val_cell = sheet.cell(
                        row=target_label_cell.row, column=target_label_cell.column + 1
                    )
                    # If empty, keep going right a few steps? (skip empty cols)
                    offset = 1
                    while (val_cell.value is None or val_cell.value == "") and offset < 5:
                        offset += 1
                        val_cell = sheet.cell(
                            row=target_label_cell.row, column=target_label_cell.column + offset
                        )

                if (val_cell is None or val_cell.value is None) and direction in ["below", "auto"]:
                    # Try below neighbor (row + 1)
                    val_cell = sheet.cell(
                        row=target_label_cell.row + 1, column=target_label_cell.column
                    )

                if val_cell and val_cell.value is not None:
                    found_val = val_cell.value
                    break  # Found a label match

            results[spec["key"]] = found_val

        return results

    def extract_table(self, start_label: str, end_label: str | None = None) -> list[dict]:
        """
        Extracts a table starting from a header row identified by start_label.
        Handles joined headers crudely but effectively.
        """
        # (Simplified implementation for this demo, would be expanded for "deep" table parsing)
        return []


if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print(json.dumps({"error": "No file provided"}))
        sys.exit(1)

    file_path = sys.argv[1]
    # Simple CLI usage: just dump active sheet cells for debugging
    # Real usage would involve passing a JSON schema config

    try:
        parser = ExcelDeepParser(file_path)
        print(f"Successfully loaded {file_path}")
    except Exception as e:
        print(json.dumps({"error": str(e)}))
