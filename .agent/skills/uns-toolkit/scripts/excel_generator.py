#!/usr/bin/env python3
"""
Excel Report Generator - Generador Profesional de Reportes Excel
================================================================
Genera documentos Excel con formato japonés profesional para:
- 賃金台帳 (Wage Ledger)
- 勤怠表 (Attendance Sheet)
- 請求書 (Invoice)
- 源泉徴収票 (Tax Withholding Certificate)
"""

import json
import argparse
from datetime import datetime, date
from decimal import Decimal

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

    # Mock classes para cuando openpyxl no está disponible
    class Font:
        def __init__(self, **kwargs):
            pass

    class PatternFill:
        def __init__(self, **kwargs):
            pass

    class Alignment:
        def __init__(self, **kwargs):
            pass

    class Border:
        def __init__(self, **kwargs):
            pass

    class Side:
        def __init__(self, **kwargs):
            pass


# ═══════════════════════════════════════════════════════════════
# ESTILOS PROFESIONALES
# ═══════════════════════════════════════════════════════════════


class ExcelStyles:
    """Estilos profesionales para documentos japoneses."""

    # Colores UNS
    COLORS = {
        "header_bg": "1F4E79",  # Azul oscuro
        "header_fg": "FFFFFF",  # Blanco
        "subheader_bg": "D6E3F8",  # Azul claro
        "total_bg": "FFF2CC",  # Amarillo claro
        "currency_bg": "E2EFDA",  # Verde claro
        "warning_bg": "FCE4D6",  # Naranja claro
        "border": "000000",  # Negro
    }

    # Fuentes
    FONTS = {
        "title": Font(name="Yu Gothic UI", size=16, bold=True),
        "header": Font(name="Yu Gothic UI", size=11, bold=True, color=COLORS["header_fg"]),
        "subheader": Font(name="Yu Gothic UI", size=10, bold=True),
        "normal": Font(name="Yu Gothic UI", size=10),
        "small": Font(name="Yu Gothic UI", size=9),
        "currency": Font(name="Yu Gothic UI", size=10),
    }

    # Bordes
    THIN_BORDER = Border(
        left=Side(style="thin", color=COLORS["border"]),
        right=Side(style="thin", color=COLORS["border"]),
        top=Side(style="thin", color=COLORS["border"]),
        bottom=Side(style="thin", color=COLORS["border"]),
    )

    THICK_BORDER = Border(
        left=Side(style="medium", color=COLORS["border"]),
        right=Side(style="medium", color=COLORS["border"]),
        top=Side(style="medium", color=COLORS["border"]),
        bottom=Side(style="medium", color=COLORS["border"]),
    )

    # Alineaciones
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center")

    # Fills
    HEADER_FILL = PatternFill(
        start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid"
    )
    SUBHEADER_FILL = PatternFill(
        start_color=COLORS["subheader_bg"], end_color=COLORS["subheader_bg"], fill_type="solid"
    )
    TOTAL_FILL = PatternFill(
        start_color=COLORS["total_bg"], end_color=COLORS["total_bg"], fill_type="solid"
    )
    CURRENCY_FILL = PatternFill(
        start_color=COLORS["currency_bg"], end_color=COLORS["currency_bg"], fill_type="solid"
    )


# ═══════════════════════════════════════════════════════════════
# GENERADORES DE REPORTES
# ═══════════════════════════════════════════════════════════════


class ChinginReportGenerator:
    """Genera 賃金台帳 (Wage Ledger) profesional."""

    def __init__(self, company_name: str = "ユニバーサル企画株式会社"):
        self.company_name = company_name
        self.styles = ExcelStyles()

    def generate(self, employees: list[dict], period: str, output_path: str) -> str:
        """
        Genera el reporte de nóminas.

        Args:
            employees: Lista de empleados con sus datos de nómina
            period: Período (ej: "2026年01月")
            output_path: Ruta del archivo de salida
        """
        if not OPENPYXL_AVAILABLE:
            return "Error: openpyxl no disponible"

        wb = Workbook()
        ws = wb.active
        ws.title = "賃金台帳"

        # ─── Encabezado ───
        ws.merge_cells("A1:L1")
        ws["A1"] = f"賃金台帳 - {period}"
        ws["A1"].font = self.styles.FONTS["title"]
        ws["A1"].alignment = self.styles.CENTER

        ws.merge_cells("A2:L2")
        ws["A2"] = self.company_name
        ws["A2"].font = self.styles.FONTS["subheader"]
        ws["A2"].alignment = self.styles.CENTER

        # ─── Headers de columnas ───
        headers = [
            ("A", "No.", 5),
            ("B", "氏名", 15),
            ("C", "時給", 10),
            ("D", "通常時間", 10),
            ("E", "残業", 8),
            ("F", "深夜", 8),
            ("G", "休日", 8),
            ("H", "総支給額", 12),
            ("I", "健康保険", 10),
            ("J", "厚生年金", 10),
            ("K", "所得税", 10),
            ("L", "差引支給額", 12),
        ]

        row = 4
        for col, header, width in headers:
            cell = ws[f"{col}{row}"]
            cell.value = header
            cell.font = self.styles.FONTS["header"]
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
            ws.column_dimensions[col].width = width

        # ─── Datos de empleados ───
        row = 5
        totals = {"gross": 0, "health": 0, "pension": 0, "tax": 0, "net": 0}

        for i, emp in enumerate(employees, 1):
            data = [
                i,
                emp.get("name", ""),
                emp.get("hourly_rate", 0),
                emp.get("regular_hours", 0),
                emp.get("overtime_hours", 0),
                emp.get("night_hours", 0),
                emp.get("holiday_hours", 0),
                emp.get("gross_pay", 0),
                emp.get("health_insurance", 0),
                emp.get("pension", 0),
                emp.get("income_tax", 0),
                emp.get("net_pay", 0),
            ]

            for col_idx, value in enumerate(data):
                cell = ws.cell(row=row, column=col_idx + 1)
                cell.value = value
                cell.font = self.styles.FONTS["normal"]
                cell.border = self.styles.THIN_BORDER

                if col_idx >= 2:  # Columnas numéricas
                    cell.alignment = self.styles.RIGHT
                    if col_idx >= 7:  # Columnas de dinero
                        cell.number_format = "¥#,##0"
                else:
                    cell.alignment = self.styles.CENTER if col_idx == 0 else self.styles.LEFT

            # Acumular totales
            totals["gross"] += emp.get("gross_pay", 0)
            totals["health"] += emp.get("health_insurance", 0)
            totals["pension"] += emp.get("pension", 0)
            totals["tax"] += emp.get("income_tax", 0)
            totals["net"] += emp.get("net_pay", 0)

            row += 1

        # ─── Fila de totales ───
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = "合計"
        ws[f"A{row}"].font = self.styles.FONTS["subheader"]
        ws[f"A{row}"].fill = self.styles.TOTAL_FILL
        ws[f"A{row}"].alignment = self.styles.CENTER
        ws[f"A{row}"].border = self.styles.THICK_BORDER

        total_values = [
            totals["gross"],
            totals["health"],
            totals["pension"],
            totals["tax"],
            totals["net"],
        ]

        for col_idx, value in enumerate(total_values, 8):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.font = self.styles.FONTS["subheader"]
            cell.fill = self.styles.TOTAL_FILL
            cell.alignment = self.styles.RIGHT
            cell.border = self.styles.THICK_BORDER
            cell.number_format = "¥#,##0"

        # ─── Pie de página ───
        row += 2
        ws[f"A{row}"] = f"作成日: {datetime.now().strftime('%Y年%m月%d日')}"
        ws[f"A{row}"].font = self.styles.FONTS["small"]

        # Guardar
        wb.save(output_path)
        return output_path


class KintaiReportGenerator:
    """Genera 勤怠表 (Attendance Sheet) profesional."""

    def __init__(self, company_name: str = "ユニバーサル企画株式会社"):
        self.company_name = company_name
        self.styles = ExcelStyles()

    def generate(self, employee: dict, year: int, month: int, output_path: str) -> str:
        """
        Genera hoja de asistencia mensual individual.
        """
        if not OPENPYXL_AVAILABLE:
            return "Error: openpyxl no disponible"

        import calendar

        wb = Workbook()
        ws = wb.active
        ws.title = "勤怠表"

        # ─── Encabezado ───
        ws.merge_cells("A1:H1")
        ws["A1"] = f"勤怠表 - {year}年{month:02d}月"
        ws["A1"].font = self.styles.FONTS["title"]
        ws["A1"].alignment = self.styles.CENTER

        # Info del empleado
        ws["A3"] = "氏名:"
        ws["B3"] = employee.get("name", "")
        ws["D3"] = "派遣先:"
        ws["E3"] = employee.get("factory", "")

        # ─── Headers de días ───
        headers = ["日付", "曜日", "出勤", "退勤", "休憩", "通常", "残業", "深夜"]
        row = 5

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.styles.FONTS["header"]
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 10

        # ─── Días del mes ───
        days_in_month = calendar.monthrange(year, month)[1]
        weekday_names = ["月", "火", "水", "木", "金", "土", "日"]

        row = 6
        totals = {"regular": 0, "overtime": 0, "night": 0}
        attendance_data = employee.get("attendance", {})

        for day in range(1, days_in_month + 1):
            d = date(year, month, day)
            weekday = weekday_names[d.weekday()]
            is_weekend = d.weekday() >= 5

            day_data = attendance_data.get(str(day), {})

            data = [
                f"{month}/{day}",
                weekday,
                day_data.get("start", ""),
                day_data.get("end", ""),
                day_data.get("break", 1.0),
                day_data.get("regular", 0),
                day_data.get("overtime", 0),
                day_data.get("night", 0),
            ]

            for col_idx, value in enumerate(data):
                cell = ws.cell(row=row, column=col_idx + 1)
                cell.value = value
                cell.font = self.styles.FONTS["normal"]
                cell.alignment = self.styles.CENTER
                cell.border = self.styles.THIN_BORDER

                # Resaltar fines de semana
                if is_weekend:
                    cell.fill = PatternFill(
                        start_color="FFF0F0", end_color="FFF0F0", fill_type="solid"
                    )

            # Acumular
            totals["regular"] += day_data.get("regular", 0)
            totals["overtime"] += day_data.get("overtime", 0)
            totals["night"] += day_data.get("night", 0)

            row += 1

        # ─── Totales ───
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "合計"
        ws[f"A{row}"].font = self.styles.FONTS["subheader"]
        ws[f"A{row}"].fill = self.styles.TOTAL_FILL
        ws[f"A{row}"].alignment = self.styles.CENTER
        ws[f"A{row}"].border = self.styles.THICK_BORDER

        for col_idx, key in enumerate(["regular", "overtime", "night"], 6):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = totals[key]
            cell.font = self.styles.FONTS["subheader"]
            cell.fill = self.styles.TOTAL_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THICK_BORDER

        # ─── Firmas ───
        row += 3
        ws[f"A{row}"] = "本人確認:"
        ws[f"D{row}"] = "承認:"

        wb.save(output_path)
        return output_path


class InvoiceGenerator:
    """Genera 請求書 (Invoice) profesional."""

    def __init__(self, company_info: dict = None):
        self.company = company_info or {
            "name": "ユニバーサル企画株式会社",
            "permit": "派23-303669",
            "address": "愛知県名古屋市港区名港二丁目6-30",
            "phone": "052-938-8840",
            "bank": "愛知銀行 当知支店 普通2075479",
        }
        self.styles = ExcelStyles()

    def generate(self, client: dict, items: list[dict], period: str, output_path: str) -> str:
        """
        Genera factura profesional.

        Args:
            client: Datos del cliente (name, address)
            items: Lista de items (description, hours, rate, amount)
            period: Período de facturación
            output_path: Ruta de salida
        """
        if not OPENPYXL_AVAILABLE:
            return "Error: openpyxl no disponible"

        wb = Workbook()
        ws = wb.active
        ws.title = "請求書"

        # Configurar ancho de columnas
        widths = [5, 30, 10, 12, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ─── Título ───
        ws.merge_cells("A1:E1")
        ws["A1"] = "請求書"
        ws["A1"].font = Font(name="Yu Gothic UI", size=24, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        # ─── Número y fecha ───
        invoice_no = datetime.now().strftime("INV-%Y%m%d-%H%M")
        ws["A3"] = f"請求書番号: {invoice_no}"
        ws["D3"] = f"発行日: {datetime.now().strftime('%Y年%m月%d日')}"

        # ─── Cliente ───
        ws["A5"] = f"{client.get('name', '')} 御中"
        ws["A5"].font = Font(name="Yu Gothic UI", size=14, bold=True)
        ws["A6"] = f"〒 {client.get('postal', '')}"
        ws["A7"] = client.get("address", "")

        # ─── Emisor ───
        ws["D5"] = self.company["name"]
        ws["D5"].font = Font(name="Yu Gothic UI", size=11, bold=True)
        ws["D6"] = f"派遣許可番号: {self.company['permit']}"
        ws["D7"] = self.company["address"]
        ws["D8"] = f"TEL: {self.company['phone']}"

        # ─── Período ───
        ws["A10"] = f"請求期間: {period}"
        ws["A10"].font = self.styles.FONTS["subheader"]

        # ─── Tabla de items ───
        headers = ["No.", "内容", "時間", "単価", "金額"]
        row = 12

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.styles.FONTS["header"]
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER

        row = 13
        subtotal = 0

        for i, item in enumerate(items, 1):
            amount = item.get("amount", item.get("hours", 0) * item.get("rate", 0))
            subtotal += amount

            data = [
                i,
                item.get("description", ""),
                item.get("hours", 0),
                item.get("rate", 0),
                amount,
            ]

            for col_idx, value in enumerate(data):
                cell = ws.cell(row=row, column=col_idx + 1)
                cell.value = value
                cell.font = self.styles.FONTS["normal"]
                cell.border = self.styles.THIN_BORDER

                if col_idx in [2, 3, 4]:
                    cell.alignment = self.styles.RIGHT
                    if col_idx >= 3:
                        cell.number_format = "¥#,##0"
                else:
                    cell.alignment = self.styles.CENTER if col_idx == 0 else self.styles.LEFT

            row += 1

        # ─── Totales ───
        tax_rate = Decimal("0.10")
        tax = int(subtotal * tax_rate)
        total = subtotal + tax

        row += 1
        totals_data = [
            ("小計", subtotal),
            ("消費税 (10%)", tax),
            ("合計金額", total),
        ]

        for label, value in totals_data:
            ws.merge_cells(f"A{row}:D{row}")
            ws[f"A{row}"] = label
            ws[f"A{row}"].alignment = self.styles.RIGHT
            ws[f"A{row}"].font = self.styles.FONTS["subheader"]

            cell = ws[f"E{row}"]
            cell.value = value
            cell.number_format = "¥#,##0"
            cell.alignment = self.styles.RIGHT
            cell.font = self.styles.FONTS["subheader"]

            if label == "合計金額":
                ws[f"A{row}"].fill = self.styles.TOTAL_FILL
                cell.fill = self.styles.TOTAL_FILL
                cell.border = self.styles.THICK_BORDER

            row += 1

        # ─── Información bancaria ───
        row += 2
        ws[f"A{row}"] = "お振込先"
        ws[f"A{row}"].font = self.styles.FONTS["subheader"]
        row += 1
        ws[f"A{row}"] = self.company["bank"]

        # ─── Notas ───
        row += 2
        ws[f"A{row}"] = "※ お支払い期限: 翌月末日"
        ws[f"A{row}"].font = self.styles.FONTS["small"]

        wb.save(output_path)
        return output_path


# ═══════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════


def main():
    parser = argparse.ArgumentParser(description="Excel Report Generator")
    parser.add_argument(
        "report_type", choices=["chingin", "kintai", "invoice"], help="Tipo de reporte"
    )
    parser.add_argument("--data", "-d", required=True, help="Archivo JSON con datos")
    parser.add_argument("--output", "-o", required=True, help="Archivo de salida")
    parser.add_argument("--period", "-p", help="Período (ej: 2026年01月)")

    args = parser.parse_args()

    # Cargar datos
    with open(args.data, encoding="utf-8") as f:
        data = json.load(f)

    period = args.period or datetime.now().strftime("%Y年%m月")

    if args.report_type == "chingin":
        generator = ChinginReportGenerator()
        result = generator.generate(data.get("employees", []), period, args.output)

    elif args.report_type == "kintai":
        generator = KintaiReportGenerator()
        year, month = datetime.now().year, datetime.now().month
        result = generator.generate(data, year, month, args.output)

    elif args.report_type == "invoice":
        generator = InvoiceGenerator()
        result = generator.generate(
            data.get("client", {}), data.get("items", []), period, args.output
        )

    print(f"✅ Reporte generado: {result}")


if __name__ == "__main__":
    main()
