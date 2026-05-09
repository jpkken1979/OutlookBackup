#!/usr/bin/env python3
"""
UNS CLI - Herramienta Unificada para Gestión de 派遣社員
=====================================================
Un solo comando para todas las operaciones de RRHH japonés.

Uso:
    uns <comando> <subcomando> [opciones]

Comandos disponibles:
    payroll   - Gestión de nóminas (賃金台帳)
    visa      - Control de visas (在留管理)
    employee  - Registro de empleados (社員台帳)
    invoice   - Facturación (請求書)
    attendance- Asistencia (勤怠)
    report    - Generación de reportes
    validate  - Validación de datos
    convert   - Conversión de formatos
    backup    - Sistema de respaldo
"""

import argparse
import json
import sys
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

# Configuración de rutas
TOOLKIT_DIR = Path(__file__).parent.parent
ARTIFACTS_DIR = Path("artifacts/uns")
BACKUP_DIR = Path("backups/uns")


# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN GLOBAL
# ═══════════════════════════════════════════════════════════════

CONFIG = {
    "company": {
        "name": "ユニバーサル企画株式会社",
        "permit": "派23-303669",
        "address": "愛知県名古屋市港区名港二丁目6-30",
        "phone": "052-938-8840",
    },
    "tax_rates_2026": {
        "health_insurance": Decimal("0.05"),
        "pension": Decimal("0.0915"),
        "employment_insurance": Decimal("0.006"),
    },
    "premium_rates": {
        "overtime": Decimal("1.25"),
        "late_night": Decimal("1.25"),
        "holiday": Decimal("1.35"),
        "overtime_60h": Decimal("1.50"),
    },
}


# ═══════════════════════════════════════════════════════════════
# VALIDADORES DE DATOS JAPONESES
# ═══════════════════════════════════════════════════════════════


class JapaneseValidators:
    """Validadores de datos específicos de Japón."""

    @staticmethod
    def validate_my_number(number: str) -> tuple[bool, str]:
        """
        Valida マイナンバー (My Number) - 12 dígitos con checksum.
        """
        # Limpiar entrada
        clean = number.replace("-", "").replace(" ", "")

        if not clean.isdigit():
            return False, "マイナンバーは数字のみです"

        if len(clean) != 12:
            return False, f"マイナンバーは12桁です (actual: {len(clean)})"

        # Validar checksum (últimos 11 dígitos -> checksum)
        digits = [int(d) for d in clean]
        weights = [6, 5, 4, 3, 2, 7, 6, 5, 4, 3, 2]

        total = sum(d * w for d, w in zip(digits[:11], weights, strict=False))
        remainder = total % 11
        check_digit = 0 if remainder <= 1 else 11 - remainder

        if digits[11] != check_digit:
            return False, "チェックディジットが無効です"

        return True, "✓ 有効なマイナンバー"

    @staticmethod
    def validate_zairyu_card(number: str) -> tuple[bool, str]:
        """
        Valida 在留カード番号 (Residence Card Number).
        Formato: AA00000000BB (2 letras + 8 números + 2 letras)
        """
        clean = number.upper().replace("-", "").replace(" ", "")

        if len(clean) != 12:
            return False, f"在留カード番号は12文字です (actual: {len(clean)})"

        # Primeros 2 caracteres: letras
        if not clean[:2].isalpha():
            return False, "最初の2文字は英字です"

        # Siguientes 8 caracteres: números
        if not clean[2:10].isdigit():
            return False, "中央の8文字は数字です"

        # Últimos 2 caracteres: letras
        if not clean[10:].isalpha():
            return False, "最後の2文字は英字です"

        return True, "✓ 有効な在留カード番号"

    @staticmethod
    def validate_phone(number: str) -> tuple[bool, str]:
        """
        Valida 電話番号 (Japanese Phone Number).
        Formatos: 03-1234-5678, 090-1234-5678, etc.
        """
        clean = number.replace("-", "").replace(" ", "").replace("(", "").replace(")", "")

        if not clean.isdigit():
            return False, "電話番号は数字のみです"

        if len(clean) < 10 or len(clean) > 11:
            return False, f"電話番号は10-11桁です (actual: {len(clean)})"

        # Validar prefijos comunes
        valid_prefixes = ["03", "06", "052", "070", "080", "090", "050"]
        has_valid_prefix = any(clean.startswith(p) for p in valid_prefixes)

        if not has_valid_prefix:
            return False, "不明な市外局番です"

        return True, "✓ 有効な電話番号"

    @staticmethod
    def validate_postal_code(code: str) -> tuple[bool, str]:
        """
        Valida 郵便番号 (Postal Code).
        Formato: 123-4567 o 1234567
        """
        clean = code.replace("-", "").replace(" ", "")

        if not clean.isdigit():
            return False, "郵便番号は数字のみです"

        if len(clean) != 7:
            return False, f"郵便番号は7桁です (actual: {len(clean)})"

        return True, "✓ 有効な郵便番号"

    @staticmethod
    def validate_bank_account(
        bank_code: str, branch_code: str, account_number: str
    ) -> tuple[bool, str]:
        """
        Valida 銀行口座 (Bank Account).
        """
        errors = []

        # Código de banco: 4 dígitos
        if not bank_code.isdigit() or len(bank_code) != 4:
            errors.append("銀行コードは4桁の数字です")

        # Código de sucursal: 3 dígitos
        if not branch_code.isdigit() or len(branch_code) != 3:
            errors.append("支店コードは3桁の数字です")

        # Número de cuenta: 7 dígitos
        if not account_number.isdigit() or len(account_number) != 7:
            errors.append("口座番号は7桁の数字です")

        if errors:
            return False, "; ".join(errors)

        return True, "✓ 有効な銀行口座"


# ═══════════════════════════════════════════════════════════════
# CALCULADORA FISCAL COMPLETA
# ═══════════════════════════════════════════════════════════════


class TaxCalculator:
    """Calculadora fiscal japonesa 2026."""

    # Brackets de impuesto sobre la renta 2026
    INCOME_TAX_BRACKETS = [
        (1_950_000, Decimal("0.05"), 0),
        (3_300_000, Decimal("0.10"), 97_500),
        (6_950_000, Decimal("0.20"), 427_500),
        (9_000_000, Decimal("0.23"), 636_000),
        (18_000_000, Decimal("0.33"), 1_536_000),
        (40_000_000, Decimal("0.40"), 2_796_000),
        (float("inf"), Decimal("0.45"), 4_796_000),
    ]

    @classmethod
    def calculate_income_tax(cls, annual_income: Decimal) -> dict:
        """
        Calcula 所得税 (Income Tax) anual.
        """
        # Deducción básica
        basic_deduction = Decimal("480000")  # 基礎控除 48万円

        # Ingreso gravable
        taxable_income = max(Decimal("0"), annual_income - basic_deduction)

        # Encontrar bracket
        tax = Decimal("0")
        for limit, rate, deduction in cls.INCOME_TAX_BRACKETS:
            if taxable_income <= limit:
                tax = (taxable_income * rate) - Decimal(str(deduction))
                break

        # Impuesto de reconstrucción (2.1%)
        reconstruction_tax = (tax * Decimal("0.021")).quantize(Decimal("1"), ROUND_HALF_UP)
        total_tax = (tax + reconstruction_tax).quantize(Decimal("1"), ROUND_HALF_UP)

        return {
            "gross_income": int(annual_income),
            "basic_deduction": int(basic_deduction),
            "taxable_income": int(taxable_income),
            "base_tax": int(tax),
            "reconstruction_tax": int(reconstruction_tax),
            "total_annual_tax": int(total_tax),
            "monthly_tax": int(total_tax / 12),
        }

    @classmethod
    def calculate_social_insurance(cls, monthly_salary: Decimal) -> dict:
        """
        Calcula 社会保険料 (Social Insurance).
        """
        rates = CONFIG["tax_rates_2026"]

        health = (monthly_salary * rates["health_insurance"]).quantize(Decimal("1"), ROUND_HALF_UP)
        pension = (monthly_salary * rates["pension"]).quantize(Decimal("1"), ROUND_HALF_UP)
        employment = (monthly_salary * rates["employment_insurance"]).quantize(
            Decimal("1"), ROUND_HALF_UP
        )

        total = health + pension + employment

        return {
            "health_insurance": int(health),
            "pension": int(pension),
            "employment_insurance": int(employment),
            "total": int(total),
            "rates": {
                "health": f"{float(rates['health_insurance']) * 100}%",
                "pension": f"{float(rates['pension']) * 100}%",
                "employment": f"{float(rates['employment_insurance']) * 100}%",
            },
        }

    @classmethod
    def calculate_resident_tax(cls, annual_income: Decimal, municipality: str = "名古屋市") -> dict:
        """
        Calcula 住民税 (Resident Tax) - aproximado.
        Nota: Varía por municipio, este es un cálculo estándar.
        """
        # Deducción básica
        basic_deduction = Decimal("430000")  # 基礎控除 43万円

        # Ingreso gravable
        taxable = max(Decimal("0"), annual_income - basic_deduction)

        # Tasa estándar: 10% (6% prefectura + 4% municipio)
        tax = (taxable * Decimal("0.10")).quantize(Decimal("1"), ROUND_HALF_UP)

        # Per cápita (均等割): aproximadamente 5,000円
        per_capita = Decimal("5000")

        total = tax + per_capita

        return {
            "taxable_income": int(taxable),
            "income_rate_tax": int(tax),
            "per_capita_tax": int(per_capita),
            "total_annual": int(total),
            "monthly": int(total / 12),
            "municipality": municipality,
        }


# ═══════════════════════════════════════════════════════════════
# CONVERSOR DE FORMATOS
# ═══════════════════════════════════════════════════════════════


class FormatConverter:
    """Conversor de formatos de archivo."""

    @staticmethod
    def excel_to_json(excel_path: str, output_path: str = None) -> dict:
        """Convierte Excel a JSON."""
        try:
            import openpyxl
        except ImportError:
            return {"error": "openpyxl no instalado. Ejecuta: pip install openpyxl"}

        wb = openpyxl.load_workbook(excel_path)
        result = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            data = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    row_dict = {h: v for h, v in zip(headers, row, strict=False) if h}
                    data.append(row_dict)

            result[sheet_name] = data

        if output_path:
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(result, f, indent=2, ensure_ascii=False, default=str)

        return result

    @staticmethod
    def json_to_csv(json_path: str, output_path: str) -> bool:
        """Convierte JSON a CSV."""
        import csv

        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)

        # Si es una lista de diccionarios
        if isinstance(data, list) and data:
            headers = list(data[0].keys())

            with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data)

            return True

        return False

    @staticmethod
    def csv_to_json(csv_path: str, output_path: str = None) -> list:
        """Convierte CSV a JSON."""
        import csv

        with open(csv_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            data = list(reader)

        if output_path:
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)

        return data


# ═══════════════════════════════════════════════════════════════
# SISTEMA DE BACKUP
# ═══════════════════════════════════════════════════════════════


class BackupSystem:
    """Sistema de backup con versionado."""

    def __init__(self, backup_dir: Path = BACKUP_DIR):
        self.backup_dir = backup_dir
        self.backup_dir.mkdir(parents=True, exist_ok=True)

    def create_backup(self, source_path: str, description: str = "") -> dict:
        """Crea un backup versionado."""
        source = Path(source_path)
        if not source.exists():
            return {"error": f"Archivo no existe: {source_path}"}

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{source.stem}_{timestamp}{source.suffix}"
        backup_path = self.backup_dir / backup_name

        # Copiar archivo
        import shutil

        shutil.copy2(source, backup_path)

        # Guardar metadata
        metadata = {
            "original": str(source),
            "backup": str(backup_path),
            "timestamp": timestamp,
            "description": description,
            "size_bytes": backup_path.stat().st_size,
        }

        meta_path = self.backup_dir / f"{backup_name}.meta.json"
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)

        return metadata

    def list_backups(self, pattern: str = "*") -> list:
        """Lista backups disponibles."""
        backups = []
        for meta_file in self.backup_dir.glob("*.meta.json"):
            with open(meta_file, encoding="utf-8") as f:
                metadata = json.load(f)
                if pattern == "*" or pattern in metadata["original"]:
                    backups.append(metadata)

        return sorted(backups, key=lambda x: x["timestamp"], reverse=True)

    def restore_backup(self, backup_path: str, target_path: str = None) -> dict:
        """Restaura un backup."""
        source = Path(backup_path)
        if not source.exists():
            return {"error": f"Backup no existe: {backup_path}"}

        target = (
            Path(target_path)
            if target_path
            else Path(source.stem.rsplit("_", 2)[0] + source.suffix)
        )

        import shutil

        shutil.copy2(source, target)

        return {
            "restored_from": str(source),
            "restored_to": str(target),
            "timestamp": datetime.now().isoformat(),
        }


# ═══════════════════════════════════════════════════════════════
# CLI PRINCIPAL
# ═══════════════════════════════════════════════════════════════


class UNSCLI:
    """CLI unificado para UNS."""

    def __init__(self):
        self.validators = JapaneseValidators()
        self.tax_calc = TaxCalculator()
        self.converter = FormatConverter()
        self.backup = BackupSystem()

    def run(self, args: list = None):
        """Ejecuta el CLI."""
        parser = argparse.ArgumentParser(
            description="UNS CLI - Herramienta Unificada para 派遣社員",
            formatter_class=argparse.RawDescriptionHelpFormatter,
            epilog="""
Ejemplos:
  uns validate my-number 123456789012
  uns tax income --annual 4000000
  uns tax social --monthly 250000
  uns convert excel-to-json data.xlsx -o output.json
  uns backup create data.xlsx --desc "Antes de actualizar"
  uns backup list
            """,
        )

        subparsers = parser.add_subparsers(dest="command", help="Comandos disponibles")

        # ─── Comando: validate ───
        validate_parser = subparsers.add_parser("validate", help="Validar datos japoneses")
        validate_sub = validate_parser.add_subparsers(dest="validate_type")

        my_number = validate_sub.add_parser("my-number", help="Validar マイナンバー")
        my_number.add_argument("number", help="Número a validar")

        zairyu = validate_sub.add_parser("zairyu", help="Validar 在留カード")
        zairyu.add_argument("number", help="Número a validar")

        phone = validate_sub.add_parser("phone", help="Validar 電話番号")
        phone.add_argument("number", help="Número a validar")

        postal = validate_sub.add_parser("postal", help="Validar 郵便番号")
        postal.add_argument("code", help="Código postal")

        # ─── Comando: tax ───
        tax_parser = subparsers.add_parser("tax", help="Cálculos fiscales")
        tax_sub = tax_parser.add_subparsers(dest="tax_type")

        income = tax_sub.add_parser("income", help="Calcular 所得税")
        income.add_argument("--annual", "-a", type=int, required=True, help="Ingreso anual en JPY")

        social = tax_sub.add_parser("social", help="Calcular 社会保険")
        social.add_argument(
            "--monthly", "-m", type=int, required=True, help="Salario mensual en JPY"
        )

        resident = tax_sub.add_parser("resident", help="Calcular 住民税")
        resident.add_argument(
            "--annual", "-a", type=int, required=True, help="Ingreso anual en JPY"
        )

        # ─── Comando: convert ───
        convert_parser = subparsers.add_parser("convert", help="Convertir formatos")
        convert_sub = convert_parser.add_subparsers(dest="convert_type")

        excel_json = convert_sub.add_parser("excel-to-json", help="Excel → JSON")
        excel_json.add_argument("input", help="Archivo Excel de entrada")
        excel_json.add_argument("-o", "--output", help="Archivo JSON de salida")

        json_csv = convert_sub.add_parser("json-to-csv", help="JSON → CSV")
        json_csv.add_argument("input", help="Archivo JSON de entrada")
        json_csv.add_argument("-o", "--output", required=True, help="Archivo CSV de salida")

        csv_json = convert_sub.add_parser("csv-to-json", help="CSV → JSON")
        csv_json.add_argument("input", help="Archivo CSV de entrada")
        csv_json.add_argument("-o", "--output", help="Archivo JSON de salida")

        # ─── Comando: backup ───
        backup_parser = subparsers.add_parser("backup", help="Sistema de backup")
        backup_sub = backup_parser.add_subparsers(dest="backup_action")

        backup_create = backup_sub.add_parser("create", help="Crear backup")
        backup_create.add_argument("file", help="Archivo a respaldar")
        backup_create.add_argument("--desc", "-d", default="", help="Descripción")

        backup_list = backup_sub.add_parser("list", help="Listar backups")
        backup_list.add_argument("--pattern", "-p", default="*", help="Filtro")

        backup_restore = backup_sub.add_parser("restore", help="Restaurar backup")
        backup_restore.add_argument("backup", help="Archivo de backup")
        backup_restore.add_argument("--target", "-t", help="Destino")

        # ─── Comando: info ───
        subparsers.add_parser("info", help="Información del sistema")

        # Parsear argumentos
        parsed = parser.parse_args(args)

        if not parsed.command:
            parser.print_help()
            return

        # Ejecutar comando
        self._execute_command(parsed)

    def _execute_command(self, args):
        """Ejecuta el comando parseado."""

        if args.command == "validate":
            self._handle_validate(args)

        elif args.command == "tax":
            self._handle_tax(args)

        elif args.command == "convert":
            self._handle_convert(args)

        elif args.command == "backup":
            self._handle_backup(args)

        elif args.command == "info":
            self._show_info()

    def _handle_validate(self, args):
        """Maneja comandos de validación."""
        if args.validate_type == "my-number":
            valid, msg = self.validators.validate_my_number(args.number)
        elif args.validate_type == "zairyu":
            valid, msg = self.validators.validate_zairyu_card(args.number)
        elif args.validate_type == "phone":
            valid, msg = self.validators.validate_phone(args.number)
        elif args.validate_type == "postal":
            valid, msg = self.validators.validate_postal_code(args.code)
        else:
            print("Especifica tipo de validación")
            return

        status = "✅" if valid else "❌"
        print(f"{status} {msg}")

    def _handle_tax(self, args):
        """Maneja comandos de cálculo fiscal."""
        if args.tax_type == "income":
            result = self.tax_calc.calculate_income_tax(Decimal(str(args.annual)))
            print("\n📊 所得税計算 (Income Tax Calculation)")
            print("=" * 50)
            print(f"  総収入:     ¥{result['gross_income']:,}")
            print(f"  基礎控除:   ¥{result['basic_deduction']:,}")
            print(f"  課税所得:   ¥{result['taxable_income']:,}")
            print(f"  基本税額:   ¥{result['base_tax']:,}")
            print(f"  復興特別税: ¥{result['reconstruction_tax']:,}")
            print("-" * 50)
            print(f"  年間所得税: ¥{result['total_annual_tax']:,}")
            print(f"  月額:       ¥{result['monthly_tax']:,}")

        elif args.tax_type == "social":
            result = self.tax_calc.calculate_social_insurance(Decimal(str(args.monthly)))
            print("\n📊 社会保険料計算 (Social Insurance)")
            print("=" * 50)
            print(f"  健康保険 ({result['rates']['health']}):  ¥{result['health_insurance']:,}")
            print(f"  厚生年金 ({result['rates']['pension']}): ¥{result['pension']:,}")
            print(
                f"  雇用保険 ({result['rates']['employment']}):  ¥{result['employment_insurance']:,}"
            )
            print("-" * 50)
            print(f"  合計:     ¥{result['total']:,}")

        elif args.tax_type == "resident":
            result = self.tax_calc.calculate_resident_tax(Decimal(str(args.annual)))
            print("\n📊 住民税計算 (Resident Tax)")
            print("=" * 50)
            print(f"  市区町村:   {result['municipality']}")
            print(f"  課税所得:   ¥{result['taxable_income']:,}")
            print(f"  所得割:     ¥{result['income_rate_tax']:,}")
            print(f"  均等割:     ¥{result['per_capita_tax']:,}")
            print("-" * 50)
            print(f"  年間住民税: ¥{result['total_annual']:,}")
            print(f"  月額:       ¥{result['monthly']:,}")

    def _handle_convert(self, args):
        """Maneja comandos de conversión."""
        if args.convert_type == "excel-to-json":
            result = self.converter.excel_to_json(args.input, args.output)
            if "error" in result:
                print(f"❌ {result['error']}")
            else:
                sheets = list(result.keys())
                total_rows = sum(len(v) for v in result.values())
                print(f"✅ Convertido: {len(sheets)} hojas, {total_rows} filas")
                if args.output:
                    print(f"   Guardado en: {args.output}")

        elif args.convert_type == "json-to-csv":
            success = self.converter.json_to_csv(args.input, args.output)
            if success:
                print(f"✅ Convertido: {args.output}")
            else:
                print("❌ Error en conversión")

        elif args.convert_type == "csv-to-json":
            result = self.converter.csv_to_json(args.input, args.output)
            print(f"✅ Convertido: {len(result)} registros")
            if args.output:
                print(f"   Guardado en: {args.output}")

    def _handle_backup(self, args):
        """Maneja comandos de backup."""
        if args.backup_action == "create":
            result = self.backup.create_backup(args.file, args.desc)
            if "error" in result:
                print(f"❌ {result['error']}")
            else:
                print(f"✅ Backup creado: {result['backup']}")
                print(f"   Tamaño: {result['size_bytes']:,} bytes")

        elif args.backup_action == "list":
            backups = self.backup.list_backups(args.pattern)
            if not backups:
                print("No hay backups disponibles")
            else:
                print(f"\n📦 Backups disponibles ({len(backups)}):")
                print("-" * 60)
                for b in backups[:10]:  # Mostrar últimos 10
                    print(f"  {b['timestamp']} | {Path(b['backup']).name}")
                    if b["description"]:
                        print(f"              └─ {b['description']}")

        elif args.backup_action == "restore":
            result = self.backup.restore_backup(args.backup, args.target)
            if "error" in result:
                print(f"❌ {result['error']}")
            else:
                print(f"✅ Restaurado: {result['restored_to']}")

    def _show_info(self):
        """Muestra información del sistema."""
        print("\n" + "=" * 60)
        print("  UNS TOOLKIT - Sistema de Gestión 派遣社員")
        print("=" * 60)
        print(f"\n  会社名: {CONFIG['company']['name']}")
        print(f"  許可番号: {CONFIG['company']['permit']}")
        print(f"  住所: {CONFIG['company']['address']}")
        print("\n  Versión: 1.0.0")
        print(f"  Python: {sys.version.split()[0]}")
        print(f"  Fecha: {datetime.now().strftime('%Y-%m-%d')}")
        print("\n  Componentes:")
        print("    ✓ Validadores de datos japoneses")
        print("    ✓ Calculadora fiscal 2026")
        print("    ✓ Conversor de formatos")
        print("    ✓ Sistema de backup")
        print("\n" + "=" * 60)


def main():
    cli = UNSCLI()
    cli.run()


if __name__ == "__main__":
    main()
