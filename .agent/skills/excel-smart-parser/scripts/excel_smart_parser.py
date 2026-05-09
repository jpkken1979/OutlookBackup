#!/usr/bin/env python3
"""
Excel Smart Parser - Parsing avanzado de archivos Excel.

Soporta:
- Celdas fusionadas
- Encabezados multi-fila
- Múltiples hojas
- Detección automática de tablas
- Normalización de tipos
- Evidencia de extracción

Version: 1.0.0
"""

from __future__ import annotations

import io
import re
import logging
from dataclasses import dataclass, field
from typing import Any
from datetime import datetime, date

# Imports opcionales con fallback
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

logging.basicConfig(level=logging.INFO, format="[%(asctime)s] [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)


# =============================================================================
# Data Classes
# =============================================================================


@dataclass
class CellInfo:
    """Información de una celda."""

    value: Any
    location: str  # "Sheet1!A1"
    is_merged: bool = False
    original_value: Any = None  # Valor antes de normalizar


@dataclass
class TableInfo:
    """Información de una tabla detectada."""

    sheet_name: str
    start_cell: str
    end_cell: str
    headers: list[str]
    rows: list[list[Any]]
    header_rows: int = 1


@dataclass
class ExtractedField:
    """Campo extraído con evidencia."""

    field_name: str
    value: Any
    value_type: str
    location: str
    raw_value: str
    confidence: float = 1.0


@dataclass
class ParseResult:
    """Resultado del parsing."""

    success: bool
    sheets: list[str] = field(default_factory=list)
    tables: list[TableInfo] = field(default_factory=list)
    extracted_fields: list[ExtractedField] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "sheets": self.sheets,
            "tables": [
                {
                    "sheet_name": t.sheet_name,
                    "start_cell": t.start_cell,
                    "end_cell": t.end_cell,
                    "headers": t.headers,
                    "row_count": len(t.rows),
                    "header_rows": t.header_rows,
                }
                for t in self.tables
            ],
            "extracted_fields": [
                {
                    "field": f.field_name,
                    "value": f.value,
                    "type": f.value_type,
                    "location": f.location,
                    "raw": f.raw_value,
                    "confidence": f.confidence,
                }
                for f in self.extracted_fields
            ],
            "warnings": self.warnings,
        }

    def get_evidence_list(self) -> list[dict[str, Any]]:
        """Retorna lista de evidencias en formato estándar."""
        return [
            {
                "field": f.field_name,
                "source": "excel",
                "location": f.location,
                "raw": f.raw_value,
                "confidence": f.confidence,
            }
            for f in self.extracted_fields
        ]


# =============================================================================
# Type Normalizers
# =============================================================================


class TypeNormalizer:
    """Normaliza tipos de datos de Excel."""

    # Epoch de Excel (1 de enero de 1900)
    EXCEL_EPOCH = datetime(1899, 12, 30)

    @staticmethod
    def normalize_date(value: Any) -> str | None:
        """Normaliza fecha a formato YYYY-MM-DD."""
        if value is None:
            return None

        # Ya es datetime
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")

        # Número de Excel (días desde epoch)
        if isinstance(value, (int, float)) and 1 < value < 100000:
            try:
                dt = TypeNormalizer.EXCEL_EPOCH + pd.Timedelta(days=value)
                return dt.strftime("%Y-%m-%d")
            except Exception:
                pass

        # String con formato de fecha
        if isinstance(value, str):
            # Intentar varios formatos
            formats = [
                "%Y-%m-%d",
                "%Y/%m/%d",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%Y年%m月%d日",
            ]
            for fmt in formats:
                try:
                    dt = datetime.strptime(value.strip(), fmt)
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue

        return str(value) if value else None

    @staticmethod
    def normalize_number(value: Any) -> float | None:
        """Normaliza número."""
        if value is None:
            return None

        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            # Limpiar formato de moneda/miles
            cleaned = re.sub(r"[¥$€,\s]", "", value)
            cleaned = cleaned.replace("円", "")

            try:
                return float(cleaned)
            except ValueError:
                pass

        return None

    @staticmethod
    def normalize_percentage(value: Any) -> float | None:
        """Normaliza porcentaje a decimal."""
        if value is None:
            return None

        if isinstance(value, str) and "%" in value:
            try:
                return float(value.replace("%", "").strip()) / 100
            except ValueError:
                pass

        if isinstance(value, (int, float)):
            # Si es menor que 1, probablemente ya es decimal
            if 0 <= value <= 1:
                return float(value)
            # Si es mayor, dividir por 100
            return float(value) / 100

        return None

    @staticmethod
    def detect_and_normalize(value: Any, type_hint: str = None) -> tuple[Any, str]:
        """
        Detecta el tipo y normaliza el valor.

        Returns:
            (valor_normalizado, tipo)
        """
        if value is None:
            return None, "null"

        # Si hay hint de tipo, usar ese
        if type_hint:
            if type_hint == "date":
                return TypeNormalizer.normalize_date(value), "date"
            elif type_hint == "number":
                return TypeNormalizer.normalize_number(value), "number"
            elif type_hint == "percentage":
                return TypeNormalizer.normalize_percentage(value), "number"

        # Detectar automáticamente
        if isinstance(value, bool):
            return value, "boolean"

        if isinstance(value, (datetime, date)):
            return TypeNormalizer.normalize_date(value), "date"

        if isinstance(value, (int, float)):
            return float(value), "number"

        if isinstance(value, str):
            value = value.strip()

            # Vacío
            if not value:
                return None, "null"

            # Porcentaje
            if "%" in value:
                normalized = TypeNormalizer.normalize_percentage(value)
                if normalized is not None:
                    return normalized, "number"

            # Fecha
            if re.match(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", value) or "年" in value:
                normalized = TypeNormalizer.normalize_date(value)
                if normalized:
                    return normalized, "date"

            # Número con formato
            if re.match(r"^[\d¥$€,.\-\s]+$", value):
                normalized = TypeNormalizer.normalize_number(value)
                if normalized is not None:
                    return normalized, "number"

            return value, "string"

        return str(value), "string"


# =============================================================================
# Excel Smart Parser
# =============================================================================


class ExcelSmartParser:
    """
    Parser avanzado de archivos Excel.

    Uso:
        parser = ExcelSmartParser()
        result = parser.parse("file.xlsx")
    """

    def __init__(self, normalize_types: bool = True):
        """
        Inicializa el parser.

        Args:
            normalize_types: Si normalizar tipos automáticamente
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl requerido: pip install openpyxl")

        self.normalize_types = normalize_types
        self.normalizer = TypeNormalizer()

    def _load_workbook(
        self, file_path: str | None = None, file_bytes: bytes | None = None
    ) -> openpyxl.Workbook:
        """Carga el workbook desde path o bytes."""
        if file_bytes:
            return load_workbook(io.BytesIO(file_bytes), data_only=True)
        elif file_path:
            return load_workbook(str(file_path), data_only=True)
        else:
            raise ValueError("Debe proporcionar file_path o file_bytes")

    def _get_merged_cell_value(self, sheet, row: int, col: int) -> tuple[Any, bool]:
        """
        Obtiene el valor de una celda, considerando merges.

        Returns:
            (valor, es_merged)
        """
        cell = sheet.cell(row=row, column=col)

        # Verificar si está en un rango fusionado
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Obtener la celda superior izquierda del rango
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                master_cell = sheet.cell(row=min_row, column=min_col)
                return master_cell.value, True

        return cell.value, False

    def _detect_table_bounds(
        self, sheet, start_row: int = 1, start_col: int = 1
    ) -> tuple[int, int, int, int]:
        """
        Detecta los límites de una tabla.

        Returns:
            (start_row, start_col, end_row, end_col)
        """
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Encontrar primera fila con datos
        data_start_row = start_row
        for row in range(start_row, max_row + 1):
            has_data = False
            for col in range(start_col, max_col + 1):
                if sheet.cell(row=row, column=col).value is not None:
                    has_data = True
                    break
            if has_data:
                data_start_row = row
                break

        # Encontrar última fila con datos
        data_end_row = max_row
        empty_rows = 0
        for row in range(data_start_row, max_row + 1):
            has_data = False
            for col in range(start_col, max_col + 1):
                if sheet.cell(row=row, column=col).value is not None:
                    has_data = True
                    break
            if not has_data:
                empty_rows += 1
                if empty_rows >= 3:  # 3 filas vacías consecutivas = fin de tabla
                    data_end_row = row - empty_rows
                    break
            else:
                empty_rows = 0

        # Encontrar columnas con datos
        data_start_col = start_col
        data_end_col = max_col

        # Verificar columnas vacías al inicio
        for col in range(start_col, max_col + 1):
            has_data = False
            for row in range(data_start_row, data_end_row + 1):
                if sheet.cell(row=row, column=col).value is not None:
                    has_data = True
                    break
            if has_data:
                data_start_col = col
                break

        # Verificar columnas vacías al final
        empty_cols = 0
        for col in range(data_start_col, max_col + 1):
            has_data = False
            for row in range(data_start_row, data_end_row + 1):
                if sheet.cell(row=row, column=col).value is not None:
                    has_data = True
                    break
            if not has_data:
                empty_cols += 1
                if empty_cols >= 3:
                    data_end_col = col - empty_cols
                    break
            else:
                empty_cols = 0
                data_end_col = col

        return data_start_row, data_start_col, data_end_row, data_end_col

    def _detect_header_rows(self, sheet, start_row: int, start_col: int, end_col: int) -> int:
        """
        Detecta cuántas filas forman el encabezado.

        Returns:
            Número de filas de encabezado
        """
        # Heurística: las filas de encabezado suelen tener:
        # - Más celdas con texto que números
        # - Celdas fusionadas
        # - Formato diferente (no implementado aquí)

        header_rows = 1

        for row in range(start_row, start_row + 5):  # Máximo 5 filas de header
            text_cells = 0
            number_cells = 0
            merged_cells = 0

            for col in range(start_col, end_col + 1):
                value, is_merged = self._get_merged_cell_value(sheet, row, col)

                if is_merged:
                    merged_cells += 1

                if value is None:
                    continue

                if isinstance(value, str):
                    text_cells += 1
                elif isinstance(value, (int, float)):
                    number_cells += 1

            # Si hay más números que texto, probablemente ya son datos
            if number_cells > text_cells and row > start_row:
                break

            # Si hay celdas fusionadas, probablemente es header
            if merged_cells > 0:
                header_rows = row - start_row + 1

        return header_rows

    def _extract_table(
        self, sheet, start_row: int, start_col: int, end_row: int, end_col: int, sheet_name: str
    ) -> TableInfo:
        """Extrae una tabla de la hoja."""
        # Detectar filas de encabezado
        header_rows_count = self._detect_header_rows(sheet, start_row, start_col, end_col)

        # Extraer headers
        headers = []
        for col in range(start_col, end_col + 1):
            header_parts = []
            for row in range(start_row, start_row + header_rows_count):
                value, _ = self._get_merged_cell_value(sheet, row, col)
                if value:
                    header_parts.append(str(value))

            header = " ".join(header_parts) if header_parts else f"Column{col}"
            headers.append(header)

        # Extraer filas de datos
        rows = []
        data_start_row = start_row + header_rows_count

        for row in range(data_start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                value, _ = self._get_merged_cell_value(sheet, row, col)

                if self.normalize_types and value is not None:
                    value, _ = self.normalizer.detect_and_normalize(value)

                row_data.append(value)

            # Solo agregar filas que tienen al menos un valor
            if any(v is not None for v in row_data):
                rows.append(row_data)

        start_cell = f"{get_column_letter(start_col)}{start_row}"
        end_cell = f"{get_column_letter(end_col)}{end_row}"

        return TableInfo(
            sheet_name=sheet_name,
            start_cell=f"{sheet_name}!{start_cell}",
            end_cell=f"{sheet_name}!{end_cell}",
            headers=headers,
            rows=rows,
            header_rows=header_rows_count,
        )

    def _match_schema_fields(
        self, tables: list[TableInfo], form_schema: dict[str, Any]
    ) -> list[ExtractedField]:
        """
        Extrae campos según el schema proporcionado.

        Args:
            tables: Tablas detectadas
            form_schema: Schema de campos a extraer

        Returns:
            Lista de campos extraídos
        """
        extracted = []

        for field_name, field_config in form_schema.items():
            aliases = field_config.get("aliases", [field_name])
            field_type = field_config.get("type", "string")

            # Buscar en todas las tablas
            for table in tables:
                # Buscar en headers
                for col_idx, header in enumerate(table.headers):
                    header_lower = header.lower()

                    # Verificar si el header coincide con algún alias
                    match = False
                    for alias in aliases:
                        if alias.lower() in header_lower or header_lower in alias.lower():
                            match = True
                            break

                    if not match:
                        continue

                    # Extraer el primer valor no nulo de esa columna
                    for row_idx, row in enumerate(table.rows):
                        value = row[col_idx] if col_idx < len(row) else None

                        if value is not None:
                            # Normalizar según tipo
                            if field_type == "date":
                                normalized_value = self.normalizer.normalize_date(value)
                            elif field_type == "number":
                                normalized_value = self.normalizer.normalize_number(value)
                            else:
                                normalized_value = str(value)

                            # Calcular ubicación
                            # row_idx es relativo a los datos (después del header)
                            actual_row = row_idx + table.header_rows + 1
                            col_letter = get_column_letter(col_idx + 1)
                            location = f"{table.sheet_name}!{col_letter}{actual_row}"

                            extracted.append(
                                ExtractedField(
                                    field_name=field_name,
                                    value=normalized_value,
                                    value_type=field_type,
                                    location=location,
                                    raw_value=str(value),
                                    confidence=1.0,
                                )
                            )
                            break

        return extracted

    def parse(
        self,
        file_path: str | None = None,
        file_bytes: bytes | None = None,
        sheet_names: list[str] | None = None,
        form_schema: dict[str, Any] | None = None,
        detect_tables: bool = True,
    ) -> ParseResult:
        """
        Parsea un archivo Excel.

        Args:
            file_path: Ruta al archivo
            file_bytes: Bytes del archivo
            sheet_names: Hojas específicas a procesar
            form_schema: Schema de campos a extraer
            detect_tables: Auto-detectar tablas

        Returns:
            ParseResult con datos extraídos
        """
        warnings = []

        try:
            # Cargar workbook
            wb = self._load_workbook(file_path, file_bytes)

            # Determinar hojas a procesar
            if sheet_names:
                sheets_to_process = [name for name in sheet_names if name in wb.sheetnames]
                missing = set(sheet_names) - set(sheets_to_process)
                if missing:
                    warnings.append(f"Hojas no encontradas: {missing}")
            else:
                sheets_to_process = wb.sheetnames

            tables = []

            # Procesar cada hoja
            for sheet_name in sheets_to_process:
                sheet = wb[sheet_name]

                if detect_tables:
                    # Detectar límites de la tabla
                    start_row, start_col, end_row, end_col = self._detect_table_bounds(sheet)

                    if end_row >= start_row and end_col >= start_col:
                        table = self._extract_table(
                            sheet, start_row, start_col, end_row, end_col, sheet_name
                        )
                        tables.append(table)
                else:
                    # Procesar toda la hoja
                    if sheet.max_row and sheet.max_column:
                        table = self._extract_table(
                            sheet, 1, 1, sheet.max_row, sheet.max_column, sheet_name
                        )
                        tables.append(table)

            # Extraer campos según schema si se proporcionó
            extracted_fields = []
            if form_schema and tables:
                extracted_fields = self._match_schema_fields(tables, form_schema)

            return ParseResult(
                success=True,
                sheets=sheets_to_process,
                tables=tables,
                extracted_fields=extracted_fields,
                warnings=warnings,
            )

        except Exception as e:
            logger.error(f"Error parseando Excel: {e}")
            return ParseResult(success=False, warnings=[f"Error: {str(e)}"])

    def parse_to_dataframes(
        self, file_path: str | None = None, file_bytes: bytes | None = None
    ) -> dict[str, pd.DataFrame]:
        """
        Parsea y retorna como DataFrames de pandas.

        Returns:
            Dict con nombre_hoja -> DataFrame
        """
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas requerido: pip install pandas")

        result = self.parse(file_path, file_bytes)

        dataframes = {}
        for table in result.tables:
            df = pd.DataFrame(table.rows, columns=table.headers)
            dataframes[table.sheet_name] = df

        return dataframes


# =============================================================================
# Convenience Functions
# =============================================================================


def parse_excel(
    file_path: str | None = None,
    file_bytes: bytes | None = None,
    form_schema: dict[str, Any] | None = None,
) -> ParseResult:
    """
    Función de conveniencia para parsear Excel.

    Args:
        file_path: Ruta al archivo
        file_bytes: Bytes del archivo
        form_schema: Schema de campos

    Returns:
        ParseResult
    """
    parser = ExcelSmartParser()
    return parser.parse(file_path=file_path, file_bytes=file_bytes, form_schema=form_schema)


def extract_fields_from_excel(file_path: str, fields: list[str]) -> dict[str, Any]:
    """
    Extrae campos específicos de un Excel.

    Args:
        file_path: Ruta al archivo
        fields: Lista de nombres de campos

    Returns:
        Dict con campo -> valor
    """
    schema = {f: {"type": "string"} for f in fields}
    parser = ExcelSmartParser()
    result = parser.parse(file_path=file_path, form_schema=schema)

    return {ef.field_name: ef.value for ef in result.extracted_fields}


# =============================================================================
# CLI
# =============================================================================


def main():
    import sys
    import argparse
    import json

    parser = argparse.ArgumentParser(description="Parser avanzado de archivos Excel")
    parser.add_argument("file", help="Ruta al archivo Excel")
    parser.add_argument("--sheets", nargs="+", help="Hojas específicas a procesar")
    parser.add_argument("--schema", help="Archivo JSON con schema de campos")
    parser.add_argument("--json", action="store_true", help="Salida en formato JSON")
    parser.add_argument("--verbose", action="store_true", help="Mostrar detalles de tablas")

    args = parser.parse_args()

    try:
        # Cargar schema si se proporciona
        form_schema = None
        if args.schema:
            with open(args.schema) as f:
                form_schema = json.load(f)

        excel_parser = ExcelSmartParser()
        result = excel_parser.parse(
            file_path=args.file, sheet_names=args.sheets, form_schema=form_schema
        )

        if args.json:
            output = result.to_dict()
            if args.verbose:
                # Incluir datos de las tablas
                for i, table in enumerate(result.tables):
                    output["tables"][i]["rows"] = table.rows
            print(json.dumps(output, indent=2, ensure_ascii=False, default=str))
        else:
            print(f"Éxito: {result.success}")
            print(f"Hojas procesadas: {result.sheets}")
            print("\n=== Tablas detectadas ===")

            for table in result.tables:
                print(f"\nTabla: {table.start_cell} -> {table.end_cell}")
                print(f"  Headers: {table.headers}")
                print(f"  Filas: {len(table.rows)}")

                if args.verbose:
                    for row in table.rows[:5]:  # Mostrar primeras 5 filas
                        print(f"    {row}")
                    if len(table.rows) > 5:
                        print(f"    ... y {len(table.rows) - 5} filas más")

            if result.extracted_fields:
                print("\n=== Campos extraídos ===")
                for ef in result.extracted_fields:
                    print(f"  {ef.field_name}: {ef.value} ({ef.location})")

            if result.warnings:
                print("\n=== Advertencias ===")
                for w in result.warnings:
                    print(f"  - {w}")

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
