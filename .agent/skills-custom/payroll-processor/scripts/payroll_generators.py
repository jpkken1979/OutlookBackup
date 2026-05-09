"""Payroll output generators — JSON, HTML, Excel.

Generates payslips in multiple formats from Payslip models.
"""

import json
import logging
from pathlib import Path
from typing import Any, Dict, Optional
from datetime import datetime

import sys

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from jinja2 import Environment, FileSystemLoader, TemplateNotFound
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from _uns_employee_schema import Payslip, Employee

logger = logging.getLogger(__name__)


class PayslipGenerator:
    """Generates payslips in JSON, HTML, and Excel formats."""

    def __init__(self, template_dir: Path | None = None) -> None:
        """Initialize generator with template directory.

        Args:
            template_dir: Path to Jinja2 templates directory.
                If None, uses ./templates relative to this file.

        Raises:
            FileNotFoundError: If template_dir doesn't exist.
        """
        if template_dir is None:
            template_dir = Path(__file__).parent.parent / "templates"

        if not template_dir.exists():
            raise FileNotFoundError(f"Template directory not found: {template_dir}")

        self.template_dir = template_dir
        self.jinja_env = Environment(loader=FileSystemLoader(str(template_dir)))
        logger.info(f"Initialized PayslipGenerator with templates at {template_dir}")

    def to_json(self, payslip: Payslip) -> dict[str, Any]:
        """Convert payslip to JSON-serializable dictionary.

        Args:
            payslip: Payslip Pydantic model.

        Returns:
            Dict[str, Any]: JSON-compatible dictionary.
        """
        data = payslip.model_dump()
        logger.debug(f"Converted payslip {payslip.id} to JSON")
        return data

    def to_html(
        self,
        payslip: Payslip,
        employee: Employee | None = None,
        template_name: str = "payslip.html",
    ) -> str:
        """Render payslip as HTML using Jinja2 template.

        Args:
            payslip: Payslip model.
            employee: Optional Employee model (for additional data).
            template_name: Name of template file.

        Returns:
            str: Rendered HTML string.

        Raises:
            TemplateNotFound: If template not found.
            jinja2.exceptions.TemplateError: If rendering fails.
        """
        try:
            template = self.jinja_env.get_template(template_name)

            context = {
                "payslip": payslip,
                "employee": employee,
                "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }

            html = template.render(context)
            logger.info(f"Rendered HTML for payslip {payslip.id}")
            return html

        except TemplateNotFound:
            logger.error(f"Template not found: {template_name}")
            raise
        except Exception as e:
            logger.error(f"HTML rendering failed for {payslip.id}: {e}")
            raise

    def to_excel(
        self, payslip: Payslip, employee: Employee | None = None, output_path: Path | None = None
    ) -> Path:
        """Export payslip to Excel file.

        Args:
            payslip: Payslip model.
            employee: Optional Employee model.
            output_path: Path to output Excel file.
                If None, creates in current directory.

        Returns:
            Path: Path to created Excel file.

        Raises:
            ValueError: If output_path parent directory doesn't exist.
        """
        if output_path is None:
            output_path = Path.cwd() / f"payslip_{payslip.id}.xlsx"

        if not output_path.parent.exists():
            raise ValueError(f"Output directory doesn't exist: {output_path.parent}")

        wb = Workbook()
        ws = wb.active
        ws.title = "給与明細"

        # Define styles
        PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        Font(bold=True, color="FFFFFF", size=12)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        # Title
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "給与明細（Payslip）"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center_align

        # Employee info section
        row = 3
        ws[f"A{row}"] = "Employee Information"
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].fill = section_fill

        row += 1
        ws[f"A{row}"] = "従業員ID / Employee ID:"
        ws[f"B{row}"] = payslip.employee_id
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        if employee:
            ws[f"A{row}"] = "氏名 / Name:"
            ws[f"B{row}"] = employee.name_jp
            ws[f"C{row}"] = employee.name_en
            row += 1

        # Period section
        ws[f"A{row}"] = "給与計算期間 / Payroll Period"
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].fill = section_fill
        row += 1

        ws[f"A{row}"] = "年月 / Year-Month:"
        ws[f"B{row}"] = f"{payslip.year}-{payslip.month:02d}"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        # Hours section
        ws[f"A{row}"] = "勤務時間 / Hours Worked"
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].fill = section_fill
        row += 1

        ws[f"A{row}"] = "通常勤務時間 / Regular Hours:"
        ws[f"B{row}"] = payslip.regular_hours
        ws[f"B{row}"].alignment = right_align
        row += 1

        ws[f"A{row}"] = "残業時間 / Overtime Hours:"
        ws[f"B{row}"] = payslip.overtime_hours
        ws[f"B{row}"].alignment = right_align
        row += 2

        # Gross pay section
        ws[f"A{row}"] = "支給額 / Earnings"
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].fill = section_fill
        row += 1

        ws[f"A{row}"] = "通常給 / Regular Pay:"
        ws[f"B{row}"] = payslip.regular_pay
        ws[f"B{row}"].number_format = "¥#,##0.00"
        ws[f"B{row}"].alignment = right_align
        row += 1

        ws[f"A{row}"] = "残業代 / Overtime Pay:"
        ws[f"B{row}"] = payslip.overtime_pay
        ws[f"B{row}"].number_format = "¥#,##0.00"
        ws[f"B{row}"].alignment = right_align
        row += 1

        ws[f"A{row}"] = "賞与 / Bonus:"
        ws[f"B{row}"] = payslip.bonus
        ws[f"B{row}"].number_format = "¥#,##0.00"
        ws[f"B{row}"].alignment = right_align
        row += 1

        ws[f"A{row}"] = "総支給額 / Gross Pay:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = payslip.gross_pay
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].number_format = "¥#,##0.00"
        ws[f"B{row}"].fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )
        ws[f"B{row}"].alignment = right_align
        row += 2

        # Deductions section
        ws[f"A{row}"] = "控除 / Deductions"
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].fill = section_fill
        row += 1

        ws[f"A{row}"] = "所得税 / Income Tax:"
        ws[f"B{row}"] = payslip.income_tax
        ws[f"B{row}"].number_format = "¥#,##0.00"
        ws[f"B{row}"].alignment = right_align
        row += 1

        ws[f"A{row}"] = "社会保険 / Social Insurance:"
        ws[f"B{row}"] = payslip.social_insurance
        ws[f"B{row}"].number_format = "¥#,##0.00"
        ws[f"B{row}"].alignment = right_align
        row += 1

        ws[f"A{row}"] = "雇用保険 / Employment Insurance:"
        ws[f"B{row}"] = payslip.employment_insurance
        ws[f"B{row}"].number_format = "¥#,##0.00"
        ws[f"B{row}"].alignment = right_align
        row += 1

        ws[f"A{row}"] = "その他控除 / Other Deductions:"
        ws[f"B{row}"] = payslip.other_deductions
        ws[f"B{row}"].number_format = "¥#,##0.00"
        ws[f"B{row}"].alignment = right_align
        row += 1

        ws[f"A{row}"] = "控除合計 / Total Deductions:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = payslip.total_deductions
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].number_format = "¥#,##0.00"
        ws[f"B{row}"].fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )
        ws[f"B{row}"].alignment = right_align
        row += 2

        # Net pay section
        ws[f"A{row}"] = "手取り額 / Net Pay"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"A{row}"].fill = PatternFill(
            start_color="92D050", end_color="92D050", fill_type="solid"
        )
        ws[f"B{row}"] = payslip.net_pay
        ws[f"B{row}"].font = Font(bold=True, size=12)
        ws[f"B{row}"].number_format = "¥#,##0.00"
        ws[f"B{row}"].fill = PatternFill(
            start_color="92D050", end_color="92D050", fill_type="solid"
        )
        ws[f"B{row}"].alignment = right_align

        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20

        # Save
        wb.save(output_path)
        logger.info(f"Saved Excel payslip to {output_path}")
        return output_path
