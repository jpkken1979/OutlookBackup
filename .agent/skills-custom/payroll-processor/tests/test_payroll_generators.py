"""Unit tests for PayslipGenerator.

Tests JSON, HTML, and Excel output generation.
"""

import pytest
import json
from pathlib import Path
from tempfile import TemporaryDirectory

import sys

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from _uns_employee_schema import Payslip
from payroll_processor.scripts.payroll_generators import PayslipGenerator


class TestPayslipGeneratorInit:
    """Test PayslipGenerator initialization."""

    def test_init_with_default_template_dir(self) -> None:
        """Test initialization with default template directory."""
        gen = PayslipGenerator()
        assert gen.template_dir.exists()
        assert gen.template_dir.name == "templates"

    def test_init_with_custom_template_dir(self) -> None:
        """Test initialization with custom template directory."""
        with TemporaryDirectory() as tmpdir:
            gen = PayslipGenerator(template_dir=Path(tmpdir))
            assert gen.template_dir == Path(tmpdir)

    def test_init_missing_template_dir(self) -> None:
        """Test initialization fails with missing template directory."""
        nonexistent = Path("/nonexistent/templates")
        with pytest.raises(FileNotFoundError):
            PayslipGenerator(template_dir=nonexistent)


class TestToJson:
    """Test JSON conversion."""

    def test_to_json_basic(self, sample_payslip: Payslip) -> None:
        """Test conversion to JSON dictionary."""
        gen = PayslipGenerator()
        json_data = gen.to_json(sample_payslip)

        assert isinstance(json_data, dict)
        assert json_data["employee_id"] == "W001"
        assert json_data["year"] == 2026
        assert json_data["month"] == 2
        assert json_data["gross_pay"] == 210000.0
        assert json_data["net_pay"] == 167475.0

    def test_to_json_serializable(self, sample_payslip: Payslip) -> None:
        """Test that JSON output is JSON-serializable."""
        gen = PayslipGenerator()
        json_data = gen.to_json(sample_payslip)

        # Should not raise
        json_str = json.dumps(json_data)
        assert isinstance(json_str, str)
        assert "W001" in json_str

    def test_to_json_contains_all_fields(self, sample_payslip: Payslip) -> None:
        """Test that JSON contains all required fields."""
        gen = PayslipGenerator()
        json_data = gen.to_json(sample_payslip)

        required_fields = [
            "id",
            "employee_id",
            "year",
            "month",
            "regular_hours",
            "overtime_hours",
            "regular_pay",
            "overtime_pay",
            "bonus",
            "gross_pay",
            "income_tax",
            "social_insurance",
            "employment_insurance",
            "total_deductions",
            "net_pay",
        ]

        for field in required_fields:
            assert field in json_data, f"Missing field: {field}"


class TestToHtml:
    """Test HTML rendering."""

    def test_to_html_renders(self, sample_payslip: Payslip) -> None:
        """Test HTML rendering with default template."""
        gen = PayslipGenerator()
        html = gen.to_html(sample_payslip)

        assert isinstance(html, str)
        assert len(html) > 0
        assert "<!DOCTYPE html>" in html or "<html" in html

    def test_to_html_contains_employee_data(self, sample_payslip: Payslip, sample_employee) -> None:
        """Test that HTML contains employee data."""
        gen = PayslipGenerator()
        html = gen.to_html(sample_payslip, employee=sample_employee)

        assert sample_payslip.employee_id in html
        assert "W001" in html
        assert sample_employee.name_jp in html or "田中" in html

    def test_to_html_contains_payslip_data(self, sample_payslip: Payslip) -> None:
        """Test that HTML contains payslip financial data."""
        gen = PayslipGenerator()
        html = gen.to_html(sample_payslip)

        # Check for key values (as strings or formatted)
        assert "2026" in html  # Year
        assert "02" in html or "2" in html  # Month
        assert "給与明細" in html or "Payslip" in html

    def test_to_html_missing_template(self, sample_payslip: Payslip) -> None:
        """Test HTML rendering with missing template."""
        with TemporaryDirectory() as tmpdir:
            gen = PayslipGenerator(template_dir=Path(tmpdir))

            with pytest.raises((FileNotFoundError, ValueError)):
                gen.to_html(sample_payslip, template_name="nonexistent.html")


class TestToExcel:
    """Test Excel export."""

    def test_to_excel_creates_file(self, sample_payslip: Payslip) -> None:
        """Test that Excel file is created."""
        with TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_payslip.xlsx"
            gen = PayslipGenerator()
            result_path = gen.to_excel(sample_payslip, output_path=output_path)

            assert result_path.exists()
            assert result_path.suffix == ".xlsx"

    def test_to_excel_file_not_empty(self, sample_payslip: Payslip) -> None:
        """Test that Excel file has content."""
        with TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_payslip.xlsx"
            gen = PayslipGenerator()
            result_path = gen.to_excel(sample_payslip, output_path=output_path)

            assert result_path.stat().st_size > 0

    def test_to_excel_contains_employee_id(self, sample_payslip: Payslip) -> None:
        """Test that Excel contains employee ID."""
        with TemporaryDirectory() as tmpdir:
            from openpyxl import load_workbook

            output_path = Path(tmpdir) / "test_payslip.xlsx"
            gen = PayslipGenerator()
            gen.to_excel(sample_payslip, output_path=output_path)

            # Read back and verify
            wb = load_workbook(output_path)
            ws = wb.active
            content_str = str([cell.value for row in ws.iter_rows() for cell in row])

            assert "W001" in content_str

    def test_to_excel_invalid_output_dir(self, sample_payslip: Payslip) -> None:
        """Test Excel export with invalid output directory."""
        nonexistent = Path("/nonexistent/dir/test.xlsx")
        gen = PayslipGenerator()

        with pytest.raises(ValueError):
            gen.to_excel(sample_payslip, output_path=nonexistent)

    def test_to_excel_default_output_path(self, sample_payslip: Payslip) -> None:
        """Test Excel export with default output path."""
        gen = PayslipGenerator()
        result_path = gen.to_excel(sample_payslip)

        try:
            assert result_path.exists()
            assert result_path.name.startswith("payslip_")
        finally:
            # Cleanup
            if result_path.exists():
                result_path.unlink()

    def test_to_excel_contains_financial_data(self, sample_payslip: Payslip) -> None:
        """Test that Excel contains all financial data."""
        with TemporaryDirectory() as tmpdir:
            from openpyxl import load_workbook

            output_path = Path(tmpdir) / "test_payslip.xlsx"
            gen = PayslipGenerator()
            gen.to_excel(sample_payslip, output_path=output_path)

            wb = load_workbook(output_path)
            ws = wb.active

            # Check for key financial values
            all_values = [cell.value for row in ws.iter_rows() for cell in row]
            all_values_str = str(all_values)

            # Gross pay should be present
            assert 210000.0 in all_values or "210000" in all_values_str

    def test_to_excel_formatting(self, sample_payslip: Payslip) -> None:
        """Test that Excel has proper formatting."""
        with TemporaryDirectory() as tmpdir:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            output_path = Path(tmpdir) / "test_payslip.xlsx"
            gen = PayslipGenerator()
            gen.to_excel(sample_payslip, output_path=output_path)

            wb = load_workbook(output_path)
            ws = wb.active

            # Check that some cells have formatting (colors)
            has_fill = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill and cell.fill.start_color:
                        has_fill = True
                        break

            # At minimum, should have some fill colors (headers, highlights)
            assert has_fill, "Excel should have some cell formatting"


class TestPayslipGeneratorIntegration:
    """Integration tests for PayslipGenerator."""

    def test_generate_all_formats(self, sample_payslip: Payslip) -> None:
        """Test generating all output formats."""
        with TemporaryDirectory() as tmpdir:
            gen = PayslipGenerator()

            # JSON
            json_data = gen.to_json(sample_payslip)
            assert isinstance(json_data, dict)

            # HTML
            html = gen.to_html(sample_payslip)
            assert isinstance(html, str)
            assert len(html) > 100

            # Excel
            output_path = Path(tmpdir) / "test.xlsx"
            excel_path = gen.to_excel(sample_payslip, output_path=output_path)
            assert excel_path.exists()

    def test_multiple_payslips(self, sample_payslip: Payslip) -> None:
        """Test generating output for multiple payslips."""
        with TemporaryDirectory() as tmpdir:
            gen = PayslipGenerator()

            # Create multiple variations
            for month in [1, 2, 3]:
                payslip = Payslip(
                    id=f"PAYSLIP-202601{month:02d}-W001",
                    employee_id="W001",
                    year=2026,
                    month=month,
                    regular_hours=160.0,
                    overtime_hours=0.0,
                    regular_pay=200000.0,
                    overtime_pay=0.0,
                    bonus=0.0,
                    gross_pay=200000.0,
                    income_tax=20000.0,
                    social_insurance=19700.0,
                    employment_insurance=800.0,
                    other_deductions=0.0,
                    total_deductions=40500.0,
                    net_pay=159500.0,
                )

                output_path = Path(tmpdir) / f"payslip_month_{month}.xlsx"
                excel_path = gen.to_excel(payslip, output_path=output_path)
                assert excel_path.exists()
