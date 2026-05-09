"""Excel workbook builder for HR Reports - WEEK 2.

PHASE 3 OPTIMIZATION: Streaming support for large datasets, chunked processing.
"""

import logging
from typing import Dict, List, Optional, Any, Generator
from datetime import date
import pandas as pd
from io import BytesIO

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.write_only import WriteOnlyWorksheet
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel export disabled")


class ExcelReportBuilder:
    """Build multi-sheet Excel workbooks for HR reports."""

    def __init__(self, title: str = "HR Report") -> None:
        """Initialize Excel builder.

        Args:
            title: Workbook title.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel export. Install: pip install openpyxl")

        self.title = title
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self.sheets: dict[str, Any] = {}
        logger.info(f"ExcelReportBuilder initialized: {title}")

    def add_sheet(
        self,
        name: str,
        data: pd.DataFrame,
        include_summary: bool = False,
        summary_data: dict[str, Any] | None = None,
    ) -> None:
        """Add a sheet to workbook.

        Args:
            name: Sheet name (max 31 chars).
            data: DataFrame with sheet data.
            include_summary: Include summary section above data.
            summary_data: Summary stats to include.
        """
        # Sanitize sheet name (max 31 chars, no invalid chars)
        safe_name = name[:31].replace("[", "").replace("]", "").replace(":", "")

        ws = self.wb.create_sheet(title=safe_name)
        self.sheets[safe_name] = ws

        row_offset = 1

        # Add summary section if requested
        if include_summary and summary_data:
            row_offset = self._add_summary_section(ws, summary_data)

        # Add data
        self._add_data_section(ws, data, start_row=row_offset)

        logger.info(f"Sheet added: {safe_name} ({len(data)} rows)")

    def _add_summary_section(self, ws: Any, summary: dict[str, Any], start_row: int = 1) -> int:
        """Add summary section to sheet.

        Args:
            ws: Worksheet.
            summary: Summary data dict.
            start_row: Starting row (1-based).

        Returns:
            Next available row for data.
        """
        row = start_row

        # Header
        ws[f"A{row}"] = "レポート概要 (Report Summary)"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"A{row}"].fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        row += 1

        # Summary items
        for key, value in summary.items():
            ws[f"A{row}"] = str(key).replace("_", " ").title()
            ws[f"B{row}"] = str(value)
            row += 1

        # Blank row before data
        row += 1
        return row

    def _add_data_section(self, ws: Any, data: pd.DataFrame, start_row: int = 1) -> None:
        """Add data section to sheet.

        Args:
            ws: Worksheet.
            data: DataFrame with data.
            start_row: Starting row (1-based).
        """
        if data.empty:
            ws[f"A{start_row}"] = "No data available"
            return

        # Write headers
        for col_idx, column in enumerate(data.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = column
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_idx, (_, row) in enumerate(data.iterrows(), start_row + 1):
            for col_idx, value in enumerate(row.values, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format numeric values
                if isinstance(value, (int, float)):
                    if isinstance(value, float):
                        cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # Auto-adjust column widths
        for col_idx, column in enumerate(data.columns, 1):
            max_length = max(
                data[column].astype(str).map(len).max(),
                len(str(column)),
            )
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Add borders
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(
            min_row=start_row,
            max_row=start_row + len(data),
            min_col=1,
            max_col=len(data.columns),
        ):
            for cell in row:
                cell.border = thin_border

    def set_properties(
        self,
        title: str | None = None,
        subject: str | None = None,
        author: str | None = None,
    ) -> None:
        """Set workbook properties.

        Args:
            title: Document title.
            subject: Document subject.
            author: Document author.
        """
        props = self.wb.properties
        if title:
            props.title = title
        if subject:
            props.subject = subject
        if author:
            props.author = author
        logger.debug("Workbook properties set")

    def save(self, file_path: str) -> None:
        """Save workbook to file.

        Args:
            file_path: Path where to save workbook.
        """
        self.wb.save(file_path)
        logger.info(f"Workbook saved: {file_path}")

    def get_bytes(self) -> bytes:
        """Get workbook as bytes.

        Returns:
            Workbook bytes.
        """
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()

    def close(self) -> None:
        """Close and cleanup workbook."""
        if self.wb:
            self.wb.close()
            logger.debug("Workbook closed")

    def add_sheet_chunked(
        self,
        name: str,
        data_generator: Generator[pd.DataFrame, None, None],
        chunk_size: int = 1000,
    ) -> None:
        """Add sheet with chunked processing (PHASE 3 optimization for large datasets).

        Args:
            name: Sheet name.
            data_generator: Generator yielding DataFrame chunks.
            chunk_size: Rows per chunk (default 1000).
        """
        safe_name = name[:31].replace("[", "").replace("]", "").replace(":", "")
        ws = self.wb.create_sheet(title=safe_name)
        self.sheets[safe_name] = ws

        row_num = 1
        header_written = False
        total_rows = 0

        for chunk_df in data_generator:
            if chunk_df.empty:
                continue

            # Write headers on first chunk
            if not header_written:
                for col_idx, column in enumerate(chunk_df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = column
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                header_written = True
                row_num = 2

            # Write chunk rows
            for _, row in chunk_df.iterrows():
                for col_idx, value in enumerate(row.values, 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.value = value
                    if isinstance(value, float):
                        cell.number_format = "0.00"
                row_num += 1
                total_rows += 1

        logger.info(f"Sheet added (chunked): {safe_name} ({total_rows} rows)")


def create_multi_report_workbook(
    reports: dict[str, pd.DataFrame],
    title: str = "HR Reports",
    author: str = "HR System",
) -> ExcelReportBuilder:
    """Create workbook with multiple report sheets.

    Args:
        reports: Dict mapping sheet names to DataFrames.
        title: Workbook title.
        author: Document author.

    Returns:
        ExcelReportBuilder with all sheets.
    """
    builder = ExcelReportBuilder(title)
    builder.set_properties(title=title, author=author)

    for sheet_name, df in reports.items():
        builder.add_sheet(sheet_name, df)

    return builder
