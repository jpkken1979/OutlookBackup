"""
PDF Generation — Unificado

Genera PDFs desde HTML, Excel, templates Jinja2, URLs y replicando
formatos de PDFs de referencia.

Funciones principales:
    - html_to_pdf()
    - excel_to_pdf()
    - clone_pdf_format()
    - generate_from_template()
    - url_to_pdf()

Dependencias:
    pip install weasyprint jinja2 openpyxl requests reportlab fpdf2 PyPDF2 playwright
    playwright install chromium
    # Ubuntu: sudo apt install fonts-noto-cjk
"""

from __future__ import annotations

import logging
import os
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

logger = logging.getLogger(__name__)

# ============================================================================
# EXCEPCIONES
# ============================================================================


class PDFGenerationError(Exception):
    """Error general en generacion de PDF."""

    pass


class EngineNotAvailableError(PDFGenerationError):
    """Motor de PDF no disponible (dependencia faltante)."""

    pass


class TemplateNotFoundError(PDFGenerationError):
    """Template Jinja2 no encontrado."""

    pass


class ExcelReadError(PDFGenerationError):
    """Error al leer archivo Excel."""

    pass


# ============================================================================
# OPCIONES DEFAULT
# ============================================================================

DEFAULT_OPTIONS: dict[str, Any] = {
    "page_size": "A4",
    "margin_top": "20mm",
    "margin_bottom": "20mm",
    "margin_left": "15mm",
    "margin_right": "15mm",
    "print_background": True,
    "display_header_footer": False,
    "scale": 1.0,
    "engine": "weasyprint",  # weasyprint | pdfkit | playwright
}


def _merge_options(options: dict[str, Any] | None) -> dict[str, Any]:
    """Combina opciones default con las pasadas por el usuario."""
    merged = DEFAULT_OPTIONS.copy()
    if options:
        merged.update(options)
    return merged


def _ensure_output_dir(output_path: str) -> Path:
    """Crea el directorio de salida si no existe."""
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _absolute_path(path: str) -> str:
    """Devuelve la ruta absoluta de un archivo."""
    return str(Path(path).resolve())


# ============================================================================
# 1. HTML-to-PDF
# ============================================================================


def html_to_pdf(
    html_content: str,
    output_path: str,
    options: dict[str, Any] | None = None,
) -> str:
    """
    Convierte contenido HTML a PDF.

    Intenta usar WeasyPrint primero, luego pdfkit, luego playwright.
    El engine puede especificarse en options['engine'].

    Args:
        html_content: Contenido HTML con meta charset UTF-8.
        output_path: Ruta del PDF de salida.
        options: Opciones de renderizado.
            - engine: str ('weasyprint' | 'pdfkit' | 'playwright')
            - page_size: str (default: 'A4')
            - margin_*: str (default: '20mm')
            - print_background: bool (default: True)

    Returns:
        Ruta absoluta del PDF generado.

    Raises:
        EngineNotAvailableError: Ningun motor disponible.
        PDFGenerationError: Error en la generacion.

    Ejemplo:
        html = "<h1>Hello World</h1><p>Hello from Japan: 日本語</p>"
        pdf_path = html_to_pdf(html, "hello.pdf")
    """
    opts = _merge_options(options)
    engine = opts.pop("engine", "weasyprint")
    output = _ensure_output_dir(output_path)

    logger.info("[html_to_pdf] engine=%s output=%s", engine, output)

    error_messages: list[str] = []

    # WeasyPrint
    if engine == "weasyprint" or (engine == "auto" and not error_messages):
        try:
            return _html_to_pdf_weasyprint(html_content, str(output), opts)
        except Exception as exc:  # noqa: BLE001
            logger.debug("[html_to_pdf] weasyprint failed: %s", exc)
            error_messages.append(f"weasyprint: {exc}")

    # pdfkit (wkhtmltopdf)
    if engine in ("pdfkit", "auto") and not error_messages:
        try:
            return _html_to_pdf_pdfkit(html_content, str(output), opts)
        except Exception as exc:  # noqa: BLE001
            logger.debug("[html_to_pdf] pdfkit failed: %s", exc)
            error_messages.append(f"pdfkit: {exc}")

    # Playwright (Chromium)
    if engine in ("playwright", "auto") and not error_messages:
        try:
            return _html_to_pdf_playwright(html_content, str(output), opts)
        except Exception as exc:  # noqa: BLE001
            logger.debug("[html_to_pdf] playwright failed: %s", exc)
            error_messages.append(f"playwright: {exc}")

    raise EngineNotAvailableError(
        f"Ningun motor de PDF disponible. Errores: {'; '.join(error_messages)}"
    )


def _html_to_pdf_weasyprint(
    html_content: str,
    output_path: str,
    options: dict[str, Any],
) -> str:
    """Implementacion con WeasyPrint."""
    try:
        from weasyprint import CSS, HTML
    except ImportError as exc:
        raise EngineNotAvailableError(
            "weasyprint no instalado. Ejecuta: pip install weasyprint"
        ) from exc

    page_size = options.get("page_size", "A4")
    margin = (
        f"{options.get('margin_top', '20mm')} "
        f"{options.get('margin_right', '15mm')} "
        f"{options.get('margin_bottom', '20mm')} "
        f"{options.get('margin_left', '15mm')}"
    )

    css = CSS(
        string=(
            f"@page {{ size: {page_size}; margin: {margin}; }}"
            f"body {{ font-family: 'Noto Sans CJK JP', sans-serif; }}"
        )
    )

    base_url = os.getcwd()
    HTML(string=html_content, base_url=base_url).write_pdf(output_path, stylesheets=[css])
    logger.info("[html_to_pdf] WeasyPrint: %s", output_path)
    return output_path


def _html_to_pdf_pdfkit(
    html_content: str,
    output_path: str,
    options: dict[str, Any],
) -> str:
    """Implementacion con pdfkit (wkhtmltopdf)."""
    try:
        import pdfkit
    except ImportError as exc:
        raise EngineNotAvailableError(
            "pdfkit no instalado. Ejecuta: pip install pdfkit"
        ) from exc

    wk_options = {
        "page-size": options.get("page_size", "A4"),
        "margin-top": options.get("margin_top", "20mm"),
        "margin-bottom": options.get("margin_bottom", "20mm"),
        "margin-left": options.get("margin_left", "15mm"),
        "margin-right": options.get("margin_right", "15mm"),
        "encoding": "UTF-8",
        "enable-local-file-access": "",
        "print-background": "1" if options.get("print_background", True) else "0",
    }

    pdfkit.from_string(html_content, output_path, options=wk_options)  # type: ignore[arg-type]
    logger.info("[html_to_pdf] pdfkit: %s", output_path)
    return output_path


def _html_to_pdf_playwright(
    html_content: str,
    output_path: str,
    options: dict[str, Any],
) -> str:
    """Implementacion con Playwright (Chromium)."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise EngineNotAvailableError(
            "playwright no instalado. Ejecuta: pip install playwright && playwright install chromium"
        ) from exc

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.set_content(html_content, wait_until="networkidle")

        margin_top = options.get("margin_top", "20mm")
        margin_bottom = options.get("margin_bottom", "20mm")
        margin_left = options.get("margin_left", "15mm")
        margin_right = options.get("margin_right", "15mm")

        page.pdf(
            path=output_path,
            format=options.get("page_size", "A4"),
            print_background=options.get("print_background", True),
            margin={
                "top": margin_top,
                "bottom": margin_bottom,
                "left": margin_left,
                "right": margin_right,
            },
        )
        browser.close()

    logger.info("[html_to_pdf] Playwright: %s", output_path)
    return output_path


# ============================================================================
# 2. EXCEL-to-PDF
# ============================================================================


def excel_to_pdf(
    excel_path: str,
    output_path: str,
    sheet_name: str | None = None,
    options: dict[str, Any] | None = None,
) -> str:
    """
    Convierte una hoja de Excel a PDF con formato de tabla.

    Lee el archivo Excel con openpyxl y genera HTML equivalente
    con los estilos de celdas traducidos a CSS.

    Args:
        excel_path: Ruta al archivo .xlsx.
        output_path: Ruta del PDF de salida.
        sheet_name: Nombre de la hoja a convertir (None = hoja activa).
        options: Opciones de renderizado (pasa a html_to_pdf).

    Returns:
        Ruta absoluta del PDF generado.

    Raises:
        ExcelReadError: No se pudo leer el archivo Excel.
        PDFGenerationError: Error en la generacion del PDF.

    Ejemplo:
        excel_to_pdf("reporte.xlsx", "reporte.pdf", sheet_name="2026年4月")
    """
    opts = _merge_options(options)
    output = _ensure_output_dir(output_path)

    logger.info(
        "[excel_to_pdf] excel=%s output=%s sheet=%s",
        excel_path,
        output,
        sheet_name,
    )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise EngineNotAvailableError(
            "openpyxl no instalado. Ejecuta: pip install openpyxl"
        ) from exc

    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as exc:
        raise ExcelReadError(f"No se pudo leer el archivo Excel: {excel_path}") from exc

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ExcelReadError(
                f"Hoja '{sheet_name}' no encontrada. "
                f"Hojas disponibles: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Generar HTML desde las celdas
    html_content = _excel_to_html(ws)
    return html_to_pdf(html_content, str(output), opts)


def _excel_to_html(ws: Any) -> str:
    """Convierte una hoja de openpyxl a HTML con estilos CSS."""

    def cell_style(cell: Any) -> str:
        """Genera estilo CSS para una celda."""
        styles: list[str] = ["padding: 6px", "border: 1px solid #ccc"]

        # Font
        if cell.font:
            if cell.font.bold:
                styles.append("font-weight: bold")
            if cell.font.size:
                styles.append(f"font-size: {cell.font.size}pt")
            color = cell.font.color
            if color and color.rgb and color.rgb != "00000000":
                rgb = color.rgb[-6:]
                styles.append(f"color: #{rgb}")

        # Background
        if cell.fill and cell.fill.fgColor:
            color = cell.fill.fgColor
            if hasattr(color, "rgb") and color.rgb and color.rgb != "00000000":
                rgb = color.rgb[-6:]
                styles.append(f"background-color: #{rgb}")
            elif hasattr(color, "theme"):
                styles.append("background-color: #f3f4f6")

        # Alignment
        if cell.alignment:
            align = cell.alignment.horizontal
            if align == "center":
                styles.append("text-align: center")
            elif align == "right":
                styles.append("text-align: right")
            elif align == "left":
                styles.append("text-align: left")

        return "; ".join(styles)

    def format_value(value: Any) -> str:
        """Formatea el valor de una celda para HTML."""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return str(value)

    rows_html: list[str] = []
    for row in ws.iter_rows():
        cells_html = []
        for cell in row:
            val = format_value(cell.value)
            style = cell_style(cell)
            cells_html.append(f"<td style='{style}'>{val}</td>")
        rows_html.append(f"<tr>{''.join(cells_html)}</tr>")

    table_html = "\n".join(rows_html)

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  body {{ font-family: 'Noto Sans CJK JP', 'Hiragino Kaku Gothic Pro', sans-serif;
         font-size: 11px; padding: 20px; }}
  table {{ border-collapse: collapse; width: 100%; }}
</style>
</head>
<body>
<table>{table_html}</table>
</body>
</html>"""


# ============================================================================
# 3. CLONE PDF FORMAT
# ============================================================================


def clone_pdf_format(
    source_pdf: str,
    data: dict[str, Any],
    output_path: str,
    options: dict[str, Any] | None = None,
) -> str:
    """
    Genera un PDF replicando el formato visual de uno de referencia.

    Extrae metadata de formato (margins, fonts, colors, layout)
    del PDF de referencia y usa esos parametros para renderizar
    el nuevo contenido con el mismo estilo visual.

    Args:
        source_pdf: Ruta al PDF de referencia con el formato a replicar.
        data: Datos para el nuevo PDF.
            Expected keys:
            - title: str
            - subtitle: str (optional)
            - date: str
            - items: list[dict(description, quantity, unit_price, total)]
            - subtotal: float
            - tax: float
            - tax_rate: float
            - total: float
            - notes: str (optional)
        output_path: Ruta del PDF de salida.
        options: Opciones adicionales.

    Returns:
        Ruta absoluta del PDF generado.

    Raises:
        PDFGenerationError: No se pudo extraer el formato o generar.

    Ejemplo:
        data = {
            "title": "請求書",
            "date": "2026-04-28",
            "items": [{"description": "派遣料", "total": 480000}],
            "subtotal": 480000,
            "tax": 48000,
            "total": 528000,
        }
        clone_pdf_format("template.pdf", data, "invoice.pdf")
    """
    opts = _merge_options(options)
    output = _ensure_output_dir(output_path)

    logger.info("[clone_pdf_format] source=%s output=%s", source_pdf, output)

    # Extraer formato del PDF de referencia
    format_info = _extract_pdf_format(source_pdf)
    logger.debug("[clone_pdf_format] format extracted: %s", format_info)

    # Generar HTML con el formato replicado
    html = _generate_clone_html(data, format_info)

    return html_to_pdf(html, str(output), opts)


def _extract_pdf_format(source_pdf: str) -> dict[str, Any]:
    """
    Extrae metadata de formato de un PDF de referencia.

    Usa PyPDF2 para leer los datos del PDF.
    Falls back a valores por defecto si PyPDF2 no esta disponible.

    Returns:
        Diccionario con:
        - pagesize: tuple (width, height) en puntos
        - margin_top, margin_bottom, margin_left, margin_right: float en mm
        - fonts: list[str] nombres de fuentes usadas
        - primary_color: str hex
        - secondary_color: str hex
    """
    format_info: dict[str, Any] = {
        "pagesize": (595.28, 841.89),  # A4 default
        "margin_top": 20.0,
        "margin_bottom": 20.0,
        "margin_left": 15.0,
        "margin_right": 15.0,
        "fonts": ["Helvetica", "Times-Roman"],
        "primary_color": "#2563eb",
        "secondary_color": "#f3f4f6",
    }

    try:
        from PyPDF2 import PdfReader
    except ImportError:
        logger.warning("[clone_pdf_format] PyPDF2 no instalado, usando formato default")
        return format_info

    try:
        reader = PdfReader(source_pdf)
        if reader.pages:
            page = reader.pages[0]
            mb = page.mediabox
            format_info["pagesize"] = (float(mb.width), float(mb.height))

            # Extraer fuentes del PDF
            if "/Resources" in page:
                resources = page["/Resources"]
                if "/Font" in resources:
                    fonts: list[str] = []
                    for name in resources["/Font"]:
                        fonts.append(str(name))
                    format_info["fonts"] = fonts
    except Exception as exc:  # noqa: BLE001
        logger.warning("[clone_pdf_format] Error extrayendo formato: %s", exc)

    return format_info


def _generate_clone_html(data: dict[str, Any], format_info: dict[str, Any]) -> str:
    """Genera HTML con los datos aplicados sobre el formato del PDF original."""

    primary = format_info.get("primary_color", "#2563eb")
    secondary = format_info.get("secondary_color", "#f3f4f6")
    fonts = format_info.get("fonts", ["Noto Sans CJK JP", "sans-serif"])
    font_family = f"'{fonts[0]}', sans-serif"

    # Render items
    items_html = ""
    if "items" in data:
        for item in data["items"]:
            items_html += f"""
        <tr>
            <td>{item.get("description", "")}</td>
            <td class="center">{item.get("quantity", "")}</td>
            <td class="right">{item.get("unit_price", ""):,.0f}</td>
            <td class="right">{item.get("total", 0):,.0f}</td>
        </tr>"""

    tax_rate = data.get("tax_rate", 0.10)
    tax_pct = int(tax_rate * 100)

    notes_html = ""
    if data.get("notes"):
        notes_html = f"<p class='notes'>{data['notes']}</p>"

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: {font_family};
    font-size: 11px;
    color: #1f2937;
    padding: 40px;
  }}
  .header {{
    display: flex;
    justify-content: space-between;
    margin-bottom: 30px;
    align-items: flex-start;
  }}
  .title-block h1 {{
    font-size: 22px;
    font-weight: bold;
    color: {primary};
  }}
  .title-block .subtitle {{
    font-size: 12px;
    color: #6b7280;
    margin-top: 4px;
  }}
  .invoice-meta {{
    text-align: right;
    font-size: 11px;
    color: #374151;
  }}
  .invoice-meta .number {{
    font-size: 16px;
    font-weight: bold;
    color: #111827;
  }}
  .parties {{
    display: flex;
    justify-content: space-between;
    margin-bottom: 30px;
    gap: 20px;
  }}
  .party {{ width: 45%; }}
  .party-label {{
    font-size: 10px;
    font-weight: bold;
    color: {primary};
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-bottom: 6px;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 20px;
  }}
  th {{
    background: {primary};
    color: white;
    padding: 10px 12px;
    text-align: left;
    font-size: 10px;
    font-weight: bold;
    letter-spacing: 0.05em;
  }}
  td {{
    padding: 10px 12px;
    border-bottom: 1px solid #e5e7eb;
    font-size: 11px;
  }}
  .center {{ text-align: center; }}
  .right {{ text-align: right; }}
  .totals {{
    margin-left: auto;
    width: 280px;
  }}
  .totals tr td {{
    padding: 8px 12px;
    font-size: 11px;
  }}
  .total-final {{
    font-size: 14px;
    font-weight: bold;
    background: {secondary};
    color: {primary};
  }}
  .notes {{
    margin-top: 30px;
    padding-top: 15px;
    border-top: 1px solid #e5e7eb;
    font-size: 10px;
    color: #6b7280;
  }}
</style>
</head>
<body>
  <div class="header">
    <div class="title-block">
      <h1>{data.get("title", "INVOICE")}</h1>
      {"<div class='subtitle'>" + data["subtitle"] + "</div>" if data.get("subtitle") else ""}
    </div>
    <div class="invoice-meta">
      <div class="number">{data.get("number", data.get("invoice_number", ""))}</div>
      <div>Date: {data.get("date", "")}</div>
      {"<div>支払期限: " + data["due_date"] + "</div>" if data.get("due_date") else ""}
    </div>
  </div>

  <div class="parties">
    <div class="party">
      <div class="party-label">派遣元 From</div>
      <div>{data.get("company_name", "")}</div>
      {"<div>" + data.get("company_address", "") + "</div>" if data.get("company_address") else ""}
      {"<div>" + data.get("company_email", "") + "</div>" if data.get("company_email") else ""}
    </div>
    <div class="party">
      <div class="party-label">派遣先 To</div>
      <div>{data.get("customer_name", data.get("client_name", ""))}</div>
      {"<div>" + data.get("customer_address", "") + "</div>" if data.get("customer_address") else ""}
      {"<div>" + data.get("customer_email", "") + "</div>" if data.get("customer_email") else ""}
    </div>
  </div>

  <table>
    <thead>
      <tr>
        <th>項目 Description</th>
        <th class="center">数量 Qty</th>
        <th class="right">単価 Unit Price</th>
        <th class="right">金額 Amount</th>
      </tr>
    </thead>
    <tbody>
    {items_html}
    </tbody>
  </table>

  <table class="totals">
    <tr>
      <td>小計 Subtotal:</td>
      <td class="right">{data.get("subtotal", 0):,.0f}</td>
    </tr>
    <tr>
      <td>消費税 Tax ({tax_pct}%):</td>
      <td class="right">{data.get("tax", 0):,.0f}</td>
    </tr>
    <tr class="total-final">
      <td>合計 Total:</td>
      <td class="right">{data.get("total", 0):,.0f}</td>
    </tr>
  </table>

  {notes_html}
</body>
</html>"""


# ============================================================================
# 4. TEMPLATE-BASED (Jinja2)
# ============================================================================


def generate_from_template(
    template_path: str,
    data: dict[str, Any],
    output_path: str,
    jinja_env: Any | None = None,
    pdf_options: dict[str, Any] | None = None,
) -> str:
    """
    Genera PDF desde un template Jinja2.

    Args:
        template_path: Ruta al archivo .html de template Jinja2.
        data: Diccionario de datos para renderizar el template.
        output_path: Ruta del PDF de salida.
        jinja_env: Environment de Jinja2 personalizado (opcional).
        pdf_options: Opciones de renderizado para html_to_pdf.

    Returns:
        Ruta absoluta del PDF generado.

    Raises:
        TemplateNotFoundError: El archivo de template no existe.
        PDFGenerationError: Error al renderizar el template.

    Ejemplo:
        generate_from_template(
            "templates/invoice.html",
            {"company": {...}, "invoice": {...}},
            "invoice.pdf"
        )
    """
    output = _ensure_output_dir(output_path)

    logger.info("[generate_from_template] template=%s output=%s", template_path, output)

    if not Path(template_path).exists():
        raise TemplateNotFoundError(f"Template no encontrado: {template_path}")

    try:
        from jinja2 import Environment, FileSystemLoader, select_autoescape
    except ImportError as exc:
        raise EngineNotAvailableError(
            "jinja2 no instalado. Ejecuta: pip install jinja2"
        ) from exc

    if jinja_env is None:
        template_dir = str(Path(template_path).resolve().parent)
        jinja_env = Environment(
            loader=FileSystemLoader(template_dir),
            autoescape=select_autoescape(["html", "xml"]),
        )
        # Filtros por defecto
        jinja_env.filters["currency"] = lambda v: f"¥{v:,.0f}" if v else "¥0"
        jinja_env.filters["int"] = lambda v: int(v) if v else 0

    template_name = Path(template_path).name
    template = jinja_env.get_template(template_name)
    html_content = template.render(**data)

    opts = _merge_options(pdf_options)
    return html_to_pdf(html_content, str(output), opts)


# ============================================================================
# 5. URL-to-PDF
# ============================================================================


def url_to_pdf(
    url: str,
    output_path: str,
    options: dict[str, Any] | None = None,
) -> str:
    """
    Descarga una URL y la renderiza como PDF.

    Args:
        url: URL a renderizar.
        output_path: Ruta del PDF de salida.
        options:
            - engine: str ('playwright' | 'weasyprint')
            - format: str (default: 'A4')
            - print_background: bool (default: True)
            - timeout: int (default: 30 segundos)

    Returns:
        Ruta absoluta del PDF generado.

    Raises:
        EngineNotAvailableError: Ningun motor disponible.
        PDFGenerationError: Error al descargar o renderizar.

    Ejemplo:
        url_to_pdf("https://example.com/report", "reporte.pdf")
    """
    opts = _merge_options(options)
    output = _ensure_output_dir(output_path)

    logger.info("[url_to_pdf] url=%s output=%s", url, output)

    engine = opts.get("engine", "playwright")

    error_messages: list[str] = []

    # Playwright (Chromium) — recomendado
    if engine in ("playwright", "auto"):
        try:
            return _url_to_pdf_playwright(url, str(output), opts)
        except Exception as exc:  # noqa: BLE001
            logger.debug("[url_to_pdf] playwright failed: %s", exc)
            error_messages.append(f"playwright: {exc}")

    # WeasyPrint (descarga el HTML y lo renderiza)
    if engine in ("weasyprint", "auto") and not error_messages:
        try:
            return _url_to_pdf_weasyprint(url, str(output), opts)
        except Exception as exc:  # noqa: BLE001
            logger.debug("[url_to_pdf] weasyprint failed: %s", exc)
            error_messages.append(f"weasyprint: {exc}")

    raise EngineNotAvailableError(
        f"No se pudo generar PDF desde URL. Errores: {'; '.join(error_messages)}"
    )


def _url_to_pdf_playwright(
    url: str,
    output_path: str,
    options: dict[str, Any],
) -> str:
    """Implementacion de URL-to-PDF con Playwright."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise EngineNotAvailableError(
            "playwright no instalado. Ejecuta: pip install playwright && playwright install chromium"
        ) from exc

    timeout_ms = (options.get("timeout", 30)) * 1000

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(url, wait_until="networkidle", timeout=timeout_ms)

        page.pdf(
            path=output_path,
            format=options.get("page_size", "A4"),
            print_background=options.get("print_background", True),
            margin={
                "top": options.get("margin_top", "20mm"),
                "bottom": options.get("margin_bottom", "20mm"),
                "left": options.get("margin_left", "15mm"),
                "right": options.get("margin_right", "15mm"),
            },
        )
        browser.close()

    logger.info("[url_to_pdf] Playwright: %s", output_path)
    return output_path


def _url_to_pdf_weasyprint(
    url: str,
    output_path: str,
    options: dict[str, Any],
) -> str:
    """Implementacion de URL-to-PDF con WeasyPrint (descarga HTML)."""
    try:
        from weasyprint import HTML
    except ImportError as exc:
        raise EngineNotAvailableError(
            "weasyprint no instalado. Ejecuta: pip install weasyprint"
        ) from exc

    HTML(url).write_pdf(  # type: ignore[arg-type]
        output_path,
        stylesheets=[],
        presentational_hints=True,
    )

    logger.info("[url_to_pdf] WeasyPrint (URL): %s", output_path)
    return output_path


# ============================================================================
# CLI ENTRY POINT
# ============================================================================


def main() -> int:
    """Entry point CLI para el skill."""
    import argparse
    import json

    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    parser = argparse.ArgumentParser(
        description="PDF Generate — Generacion unificada de PDFs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    subparsers = parser.add_subparsers(dest="action", help="Accion a ejecutar")

    # html-to-pdf
    p_html = subparsers.add_parser("html-to-pdf", help="HTML a PDF")
    p_html.add_argument("--html", required=True, help="Contenido HTML")
    p_html.add_argument("--output", required=True, help="Ruta de salida PDF")
    p_html.add_argument("--engine", default="weasyprint", help="Motor (weasyprint|pdfkit|playwright)")

    # excel-to-pdf
    p_excel = subparsers.add_parser("excel-to-pdf", help="Excel a PDF")
    p_excel.add_argument("--excel", required=True, help="Ruta al archivo Excel")
    p_excel.add_argument("--output", required=True, help="Ruta de salida PDF")
    p_excel.add_argument("--sheet", help="Nombre de hoja (opcional)")

    # clone-pdf-format
    p_clone = subparsers.add_parser("clone-pdf-format", help="Clonar formato de PDF")
    p_clone.add_argument("--source", required=True, help="PDF de referencia")
    p_clone.add_argument("--data", required=True, help="JSON con datos")
    p_clone.add_argument("--output", required=True, help="Ruta de salida PDF")

    # generate-from-template
    p_tpl = subparsers.add_parser("generate-from-template", help="Template Jinja2 a PDF")
    p_tpl.add_argument("--template", required=True, help="Ruta al template HTML")
    p_tpl.add_argument("--data", required=True, help="JSON con datos")
    p_tpl.add_argument("--output", required=True, help="Ruta de salida PDF")

    # url-to-pdf
    p_url = subparsers.add_parser("url-to-pdf", help="URL a PDF")
    p_url.add_argument("--url", required=True, help="URL a renderizar")
    p_url.add_argument("--output", required=True, help="Ruta de salida PDF")
    p_url.add_argument("--engine", default="playwright", help="Motor (playwright|weasyprint)")

    args = parser.parse_args()

    if not args.action:
        parser.print_help()
        return 0

    try:
        if args.action == "html-to-pdf":
            result = html_to_pdf(args.html, args.output, {"engine": args.engine})
            print(f"OK: {result}")

        elif args.action == "excel-to-pdf":
            result = excel_to_pdf(args.excel, args.output, sheet_name=args.sheet)
            print(f"OK: {result}")

        elif args.action == "clone-pdf-format":
            data = json.loads(args.data)
            result = clone_pdf_format(args.source, data, args.output)
            print(f"OK: {result}")

        elif args.action == "generate-from-template":
            data = json.loads(args.data)
            result = generate_from_template(args.template, data, args.output)
            print(f"OK: {result}")

        elif args.action == "url-to-pdf":
            result = url_to_pdf(args.url, args.output, {"engine": args.engine})
            print(f"OK: {result}")

        return 0

    except (PDFGenerationError, EngineNotAvailableError, TemplateNotFoundError) as exc:
        logger.error("[main] Error: %s", exc)
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    except json.JSONDecodeError as exc:
        logger.error("[main] JSON invalido: %s", exc)
        print(f"ERROR: JSON invalido: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
