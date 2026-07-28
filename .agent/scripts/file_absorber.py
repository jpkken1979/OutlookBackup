#!/usr/bin/env python3
"""
File Absorber — Parser de Excel/PDF/CSV que extrae datos al Brain.

Soporta:
  - Excel (.xlsx, .xls) - hojas, columnas, rangos
  - CSV (.csv)
  - PDF (.pdf) - texto extraido
  - Markdown (.md)
  - JSON (.json)

Uso:
  python file_absorber.py <ruta>                    # absorber un archivo
  python file_absorber.py <ruta> --area business    # especificar area
  python file_absorber.py --dir <dir>               # absorber todos los archivos
  python file_absorber.py --scan-vault              # scan Obsidian vault
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

AGENT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(AGENT_ROOT))

logging.basicConfig(level=logging.INFO, format="[%(asctime)s] %(message)s")
logger = logging.getLogger(__name__)


def parse_excel(path: Path) -> list[dict]:
    """Parsea Excel y devuelve una lista de nodos por hoja."""
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError:
        logger.warning("openpyxl no instalado. pip install openpyxl")
        return []

    nodes = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        logger.error("Error cargando Excel %s: %s", path, e)
        return []

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True, max_row=200))
            if not rows:
                continue

            # Header = primera fila no vacia
            header = None
            for row in rows[:5]:
                if any(cell is not None for cell in row):
                    header = [str(c) if c is not None else "" for c in row]
                    break
            if not header:
                continue

            # Muestra de datos (primeras 20 filas post-header)
            sample = rows[1:21] if len(rows) > 1 else []
            sample_str = "\n".join(
                [
                    " | ".join([str(c)[:40] if c is not None else "-" for c in r])
                    for r in sample
                    if any(c is not None for c in r)
                ]
            )

            nodes.append(
                {
                    "title": f"Excel {path.stem} — hoja {sheet_name}",
                    "context": (
                        f"Archivo: {path.name}\n"
                        f"Hoja: {sheet_name}\n"
                        f"Filas: {len(rows)}\n\n"
                        f"Columnas ({len(header)}):\n{' | '.join(header)}\n\n"
                        f"Muestra:\n{sample_str[:1500]}"
                    ),
                    "area": "business",
                    "tags": [
                        "excel",
                        "file-absorb",
                        path.stem.lower()[:20],
                        sheet_name.lower()[:15],
                    ],
                    "node_type": "entity",
                    "importance": "high",
                }
            )
        except Exception as e:
            logger.warning("Error en hoja %s: %s", sheet_name, e)

    wb.close()
    return nodes


def parse_csv(path: Path) -> list[dict]:
    """Parsea CSV."""
    import csv  # noqa: PLC0415

    try:
        with path.open(encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = [r for _, r in zip(range(200), reader, strict=False)]
    except Exception as e:
        logger.error("Error CSV %s: %s", path, e)
        return []

    if not rows:
        return []

    header = rows[0]
    sample = rows[1:21]
    sample_str = "\n".join([" | ".join(r[:10]) for r in sample[:15]])

    return [
        {
            "title": f"CSV {path.stem}",
            "context": (
                f"Archivo: {path.name}\n"
                f"Filas: {len(rows)}\n\n"
                f"Columnas ({len(header)}):\n{' | '.join(header[:20])}\n\n"
                f"Muestra:\n{sample_str[:1500]}"
            ),
            "area": "business",
            "tags": ["csv", "file-absorb", path.stem.lower()[:20]],
            "node_type": "entity",
            "importance": "normal",
        }
    ]


def parse_pdf(path: Path) -> list[dict]:
    """Parsea PDF y extrae texto."""
    try:
        from pypdf import PdfReader  # noqa: PLC0415
    except ImportError:
        try:
            from PyPDF2 import PdfReader  # type: ignore  # noqa: PLC0415
        except ImportError:
            logger.warning("pypdf no instalado. pip install pypdf")
            return []

    try:
        reader = PdfReader(str(path))
    except Exception as e:
        logger.error("Error PDF %s: %s", path, e)
        return []

    text_parts = []
    for i, page in enumerate(reader.pages[:20]):  # max 20 paginas
        try:
            text = page.extract_text() or ""
            if text.strip():
                text_parts.append(f"=== Pagina {i + 1} ===\n{text[:1000]}")
        except Exception:
            continue

    if not text_parts:
        return []

    return [
        {
            "title": f"PDF {path.stem}",
            "context": (
                f"Archivo: {path.name}\n"
                f"Paginas: {len(reader.pages)}\n\n" + "\n\n".join(text_parts[:5])
            )[:3000],
            "area": "docs",
            "tags": ["pdf", "file-absorb", path.stem.lower()[:20]],
            "node_type": "entity",
            "importance": "normal",
        }
    ]


def parse_markdown(path: Path) -> list[dict]:
    """Parsea markdown."""
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return []

    return [
        {
            "title": f"MD {path.stem}",
            "context": f"Archivo: {path.relative_to(path.anchor)}\n\n{text[:2500]}",
            "area": "docs",
            "tags": ["markdown", "file-absorb", path.stem.lower()[:20]],
            "node_type": "session",
            "importance": "normal",
        }
    ]


def parse_json(path: Path) -> list[dict]:
    """Parsea JSON mostrando estructura."""
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []

    if isinstance(data, list):
        count = len(data)
        sample = data[:3]
        ctx = f"JSON array: {count} items\n\nMuestra:\n{json.dumps(sample, ensure_ascii=False, indent=2)[:1500]}"
    else:
        keys = list(data.keys()) if isinstance(data, dict) else []
        ctx = f"JSON object\nKeys: {keys[:20]}\n\n{json.dumps(data, ensure_ascii=False, indent=2)[:1500]}"

    return [
        {
            "title": f"JSON {path.stem}",
            "context": f"Archivo: {path.name}\n\n{ctx}",
            "area": "dev",
            "tags": ["json", "file-absorb", path.stem.lower()[:20]],
            "node_type": "entity",
            "importance": "normal",
        }
    ]


PARSERS = {
    ".xlsx": parse_excel,
    ".xls": parse_excel,
    ".csv": parse_csv,
    ".pdf": parse_pdf,
    ".md": parse_markdown,
    ".json": parse_json,
}


def absorb_file(path: Path) -> int:
    """Absorbe un archivo al Brain. Retorna numero de nodos creados."""
    ext = path.suffix.lower()
    parser = PARSERS.get(ext)
    if not parser:
        logger.debug("Skip %s (no parser para %s)", path.name, ext)
        return 0

    nodes = parser(path)
    if not nodes:
        return 0

    try:
        from core.brain import Brain  # noqa: PLC0415

        brain = Brain(AGENT_ROOT / "brain", app_id="nexus-mother")
        for node_data in nodes:
            brain.ingest(**node_data)
        logger.info("Absorbido %s: %d nodos", path.name, len(nodes))
        return len(nodes)
    except Exception as e:
        logger.error("No se pudo ingestar %s: %s", path.name, e)
        return 0


def absorb_dir(dir_path: Path, recursive: bool = True) -> int:
    """Absorbe todos los archivos soportados en un directorio."""
    total = 0
    glob = dir_path.rglob("*") if recursive else dir_path.glob("*")
    for file in glob:
        if not file.is_file():
            continue
        if any(p.startswith(".") for p in file.parts):
            continue
        if file.suffix.lower() in PARSERS:
            total += absorb_file(file)
    logger.info("Total: %d nodos en %s", total, dir_path)
    return total


def main() -> int:
    parser = argparse.ArgumentParser(description="File Absorber -> Brain")
    parser.add_argument("path", nargs="?", help="Archivo o directorio")
    parser.add_argument("--dir", help="Directorio a scannear")
    parser.add_argument("--scan-vault", action="store_true", help="Scan Obsidian vault")
    parser.add_argument("--area", default=None)
    args = parser.parse_args()

    if args.scan_vault:
        import os  # noqa: PLC0415

        vault = os.environ.get("OBSIDIAN_VAULT_ROOT", "")
        if not vault:
            vault = str(Path.home() / "Documents" / "GitHub" / "Obsidian")
        vault_path = Path(vault)
        if not vault_path.exists():
            logger.error("Vault no encontrado: %s", vault_path)
            return 1
        logger.info("Scannear vault: %s", vault_path)
        absorb_dir(vault_path, recursive=True)
        return 0

    if args.dir:
        absorb_dir(Path(args.dir))
        return 0

    if args.path:
        p = Path(args.path)
        if p.is_dir():
            absorb_dir(p)
        else:
            absorb_file(p)
        return 0

    parser.print_help()
    return 1


if __name__ == "__main__":
    sys.exit(main())
